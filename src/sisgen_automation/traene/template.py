from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from sisgen_automation.g7.catalog import load_g7_catalog, load_g7_catalog_from_db



def _load_g7_template_catalog(
    *,
    catalog_path: Path | None,
    catalog_db_path: Path | None,
):
    if catalog_db_path is not None:
        return load_g7_catalog_from_db(catalog_db_path)

    if catalog_path is not None:
        return load_g7_catalog(catalog_path)

    raise ValueError("Debe indicar catalog_path o catalog_db_path para G7.")


TRAENE_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODEMP",
    "CCODGEN",
    "CNOMGEN",
    "NTRAPOT",
]

EDITABLE_HEADERS = {"NTRAPOT"}
NUMERIC_HEADERS = {"NTRAPOT"}


@dataclass(frozen=True)
class TraeneTemplateResult:
    period: str
    output_path: Path
    rows: int


def parse_period(period: str) -> tuple[str, str]:
    try:
        year_text, month_text = period.split("-", maxsplit=1)
    except ValueError as exc:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.") from exc

    if len(year_text) != 4 or len(month_text) != 2:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.")

    year = int(year_text)
    month = int(month_text)

    if not 1 <= month <= 12:
        raise ValueError("El mes del periodo debe estar entre 01 y 12.")

    return f"{year % 100:02d}", f"{month:02d}"


def default_traene_template_path(period: str) -> Path:
    year_text, month_text = period.split("-", maxsplit=1)
    return Path("reports") / "templates" / f"TRAENE_{year_text}_{month_text}_template.xlsx"


def create_traene_template(
    *,
    period: str,
    catalog_path: Path | None = None,
    catalog_db_path: Path | None = None,
    output_path: Path | None = None,
) -> TraeneTemplateResult:
    year_short, month = parse_period(period)
    catalog = _load_g7_template_catalog(
        catalog_path=catalog_path,
        catalog_db_path=catalog_db_path,
    )

    if output_path is None:
        output_path = default_traene_template_path(period)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TRAENE"

    header_fill = PatternFill("solid", fgColor="7030A0")
    locked_fill = PatternFill("solid", fgColor="DDEBF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    for col_index, header in enumerate(TRAENE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

    for row_index, party in enumerate(catalog.power_transfers, start=2):
        values = {
            "CANOREG": year_short,
            "CMESREG": month,
            "CCODEMP": catalog.company.ccodeemp,
            "CCODGEN": party.ccodgen,
            "CNOMGEN": party.cnomgen,
            "NTRAPOT": None,
        }

        for col_index, header in enumerate(TRAENE_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=values[header])

            if header in EDITABLE_HEADERS:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

            if header in NUMERIC_HEADERS:
                cell.number_format = "0.00000"

    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 14,
        "E": 45,
        "F": 14,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{1 + len(catalog.power_transfers)}"
    sheet.sheet_view.showGridLines = False
    sheet.protection.sheet = True
    sheet.protection.enable()

    workbook.save(output_path)

    return TraeneTemplateResult(
        period=period,
        output_path=output_path,
        rows=len(catalog.power_transfers),
    )