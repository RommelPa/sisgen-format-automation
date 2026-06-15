from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

from dbfread import DBF  # type: ignore[import-untyped]
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from sisgen_automation.cenhid.template import parse_period

from sisgen_automation.cenhid.catalog import load_cenhid_catalog
from sisgen_automation.center.catalog import load_center_catalog
from sisgen_automation.catalogs.g1_repository import list_g1_units

DACOCE_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODEMP",
    "CCODCON",
    "CTIPCEN",
    "NCONPRO",
    "NPRONET",
    "NMAXDEM",
]

EDITABLE_COLUMNS = {"F", "G", "H"}
CENTER_TYPES = {"H", "T"}


@dataclass(frozen=True)
class DacoceTemplateResult:
    output_path: Path
    period: str
    base_period: str
    row_count: int


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _clean_upper(value: object) -> str:
    return _clean_text(value).upper()


def _record_period(record: Mapping[str, object]) -> str | None:
    year = _clean_text(record.get("CANOREG"))
    month = _clean_text(record.get("CMESREG"))

    if not year.isdigit() or len(year) != 2:
        return None

    if month not in {f"{item:02d}" for item in range(1, 13)}:
        return None

    return f"20{year}-{month}"


def _infer_latest_period_before_target(source_dbf_path: Path, target_period: str) -> str:
    parse_period(target_period)

    table = DBF(str(source_dbf_path), load=True, char_decode_errors="ignore")
    periods: set[str] = set()

    for record in table:
        period = _record_period(record)

        if period is None:
            continue

        if period < target_period:
            periods.add(period)

    if not periods:
        raise ValueError(
            "No se encontraron periodos válidos anteriores al periodo objetivo en DACOCE."
        )

    return sorted(periods)[-1]


def _read_base_rows(source_dbf_path: Path, base_period: str) -> list[dict[str, str]]:
    year, month = parse_period(base_period)
    table = DBF(str(source_dbf_path), load=True, char_decode_errors="ignore")

    rows: list[dict[str, str]] = []
    seen_keys: set[tuple[str, str]] = set()

    for record in table:
        if _clean_text(record.get("CANOREG")) != year:
            continue

        if _clean_text(record.get("CMESREG")) != month:
            continue

        ccodcon = _clean_text(record.get("CCODCON"))
        ctipcen = _clean_upper(record.get("CTIPCEN"))

        if not ccodcon:
            continue

        if ctipcen not in CENTER_TYPES:
            continue

        key = (ccodcon, ctipcen)

        if key in seen_keys:
            continue

        seen_keys.add(key)

        rows.append(
            {
                "CCODCON": ccodcon,
                "CTIPCEN": ctipcen,
            }
        )

    if not rows:
        raise ValueError(f"No se encontraron filas DACOCE válidas para {base_period}.")

    return rows

def _read_rows_from_catalogs(
    cenhid_catalog_path: Path,
    center_catalog_path: Path,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen_keys: set[tuple[str, str]] = set()

    cenhid_catalog = load_cenhid_catalog(cenhid_catalog_path)

    for unit in cenhid_catalog.values():
        key = (unit.ccodcon, "H")

        if key in seen_keys:
            continue

        seen_keys.add(key)
        rows.append(
            {
                "CCODCON": unit.ccodcon,
                "CTIPCEN": "H",
            }
        )

    center_catalog = load_center_catalog(center_catalog_path)

    for unit in center_catalog.values():
        key = (unit.ccodcon, "T")

        if key in seen_keys:
            continue

        seen_keys.add(key)
        rows.append(
            {
                "CCODCON": unit.ccodcon,
                "CTIPCEN": "T",
            }
        )

    return sorted(rows, key=lambda item: (item["CTIPCEN"], item["CCODCON"]))


def _read_rows_from_catalog_db(catalog_db_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen_keys: set[tuple[str, str]] = set()

    for unit in list_g1_units(
        catalog_db_path,
        active_only=True,
        visible_only=True,
    ):
        source_format = str(unit["source_format"]).upper()

        if source_format == "CENHID":
            ctipcen = "H"
        elif source_format == "CENTER":
            ctipcen = "T"
        else:
            continue

        key = (str(unit["ccodcon"]), ctipcen)

        if key in seen_keys:
            continue

        seen_keys.add(key)
        rows.append(
            {
                "CCODCON": str(unit["ccodcon"]),
                "CTIPCEN": ctipcen,
            }
        )

    if not rows:
        raise ValueError("El catalogo SQLite G1 no contiene unidades activas y visibles.")

    return sorted(rows, key=lambda item: (item["CTIPCEN"], item["CCODCON"]))


def _write_instructions_sheet(workbook: Workbook, period: str, base_period: str) -> None:
    sheet = workbook.create_sheet("INSTRUCCIONES")

    rows = [
        ("Plantilla", "DACOCE"),
        ("Periodo", period),
        ("Periodo base", base_period),
        ("Uso", "Completar solo las columnas editables."),
        ("Columnas editables", "NCONPRO, NPRONET, NMAXDEM."),
        ("NCONPRO", "Consumo propio."),
        ("NPRONET", "Producción neta. Fuente oficial para G1."),
        ("NMAXDEM", "Máxima demanda."),
        ("Advertencia", "No modificar año, mes, empresa, código ni tipo de central."),
    ]

    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90

    for cell in sheet["A"]:
        cell.font = Font(bold=True)


def _write_header(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="7030A0")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(DACOCE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(DACOCE_HEADERS))}1"


def _write_data_row(
    sheet,
    row_number: int,
    year: str,
    month: str,
    row: dict[str, str],
) -> None:
    values = [
        year,
        month,
        "EGASA",
        row["CCODCON"],
        row["CTIPCEN"],
        None,
        None,
        None,
    ]

    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_number, column=column_index, value=value)
        column_letter = get_column_letter(column_index)

        if column_letter in EDITABLE_COLUMNS:
            cell.protection = Protection(locked=False)
            cell.number_format = "0.00000"
        else:
            cell.protection = Protection(locked=True)


def _apply_format(sheet, last_row: int) -> None:
    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 14,
        "E": 10,
        "F": 16,
        "G": 16,
        "H": 16,
    }

    fixed_fill = PatternFill("solid", fgColor="D9EAF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    for row in sheet.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            column_letter = get_column_letter(cell.column)

            if column_letter in EDITABLE_COLUMNS:
                cell.fill = editable_fill
            else:
                cell.fill = fixed_fill

            cell.alignment = Alignment(horizontal="center", vertical="center")

    center_type_validation = DataValidation(
        type="list",
        formula1='"H,T"',
        allow_blank=False,
    )
    center_type_validation.error = "Solo se permite H o T."
    center_type_validation.errorTitle = "Tipo de central inválido"

    sheet.add_data_validation(center_type_validation)
    center_type_validation.add(f"E2:E{last_row}")

    sheet["E1"].comment = Comment("H = hidroeléctrica, T = térmica.", "SISGEN")
    sheet["F1"].comment = Comment("Consumo propio.", "SISGEN")
    sheet["G1"].comment = Comment("Producción neta.", "SISGEN")
    sheet["H1"].comment = Comment("Máxima demanda.", "SISGEN")

    sheet.protection.sheet = True
    sheet.protection.password = "sisgen"


def create_dacoce_template(
    period: str,
    source_dbf_path: Path | None = None,
    base_period: str | None = None,
    output_path: Path | None = None,
    cenhid_catalog_path: Path | None = None,
    center_catalog_path: Path | None = None,
    catalog_db_path: Path | None = None,
) -> DacoceTemplateResult:
    year, month = parse_period(period)

    use_catalogs = cenhid_catalog_path is not None or center_catalog_path is not None

    if catalog_db_path is not None and use_catalogs:
        raise ValueError("Use solo un origen de catalogo: YAML o SQLite.")

    if catalog_db_path is not None:
        base_period = "catalogo SQLite local"
        base_rows = _read_rows_from_catalog_db(catalog_db_path)
    elif use_catalogs:
        if cenhid_catalog_path is None or center_catalog_path is None:
            raise ValueError(
                "Para generar DACOCE desde catálogos debes indicar "
                "cenhid_catalog_path y center_catalog_path."
            )

        base_period = "catálogos locales"
        base_rows = _read_rows_from_catalogs(
            cenhid_catalog_path=cenhid_catalog_path,
            center_catalog_path=center_catalog_path,
        )
    else:
        if source_dbf_path is None:
            raise ValueError(
                "Debes indicar source_dbf_path o usar catálogos CENHID/CENTER."
            )

        if base_period is None:
            base_period = _infer_latest_period_before_target(
                source_dbf_path=source_dbf_path,
                target_period=period,
            )
        else:
            parse_period(base_period)

            if base_period >= period:
                raise ValueError(
                    "El periodo base debe ser anterior al periodo de la plantilla."
                )

        base_rows = _read_base_rows(
            source_dbf_path=source_dbf_path,
            base_period=base_period,
        )
    
    year_text, month_text = period.split("-", maxsplit=1)
    if output_path is None:
        output_path = Path("reports") / "templates" / f"DACOCE_{year_text}_{month_text}_template.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DACOCE"

    _write_header(sheet)

    for row_number, row in enumerate(base_rows, start=2):
        _write_data_row(
            sheet=sheet,
            row_number=row_number,
            year=year,
            month=month,
            row=row,
        )

    last_row = len(base_rows) + 1
    _apply_format(sheet, last_row=last_row)
    _write_instructions_sheet(workbook, period=period, base_period=base_period)

    workbook.save(output_path)

    return DacoceTemplateResult(
        output_path=output_path,
        period=period,
        base_period=base_period,
        row_count=len(base_rows),
    )