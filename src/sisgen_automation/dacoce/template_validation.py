from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from sisgen_automation.cenhid.template import parse_period
from sisgen_automation.dacoce.template import DACOCE_HEADERS

IssueSeverity = Literal["ERROR", "WARNING"]

CENTER_TYPES = {"H", "T"}
REQUIRED_NON_NEGATIVE_FIELDS = ["NCONPRO", "NMAXDEM"]
REQUIRED_SIGNED_FIELDS = ["NPRONET"]
NUMERIC_FIELDS = REQUIRED_NON_NEGATIVE_FIELDS + REQUIRED_SIGNED_FIELDS


@dataclass(frozen=True)
class DacoceTemplateIssue:
    severity: IssueSeverity
    row_number: int | None
    field: str | None
    message: str
    value: str | None = None


@dataclass(frozen=True)
class DacoceTemplateValidationResult:
    template_path: Path
    period: str
    rows_read: int
    hydro_rows: int
    thermal_rows: int
    issues: list[DacoceTemplateIssue]

    @property
    def errors(self) -> list[DacoceTemplateIssue]:
        return [issue for issue in self.issues if issue.severity == "ERROR"]

    @property
    def warnings(self) -> list[DacoceTemplateIssue]:
        return [issue for issue in self.issues if issue.severity == "WARNING"]

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


def _add_issue(
    issues: list[DacoceTemplateIssue],
    severity: IssueSeverity,
    row_number: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        DacoceTemplateIssue(
            severity=severity,
            row_number=row_number,
            field=field,
            message=message,
            value=None if value is None else str(value),
        )
    )


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _clean_upper(value: object) -> str:
    return _clean_text(value).upper()


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _validate_headers(headers: list[str], issues: list[DacoceTemplateIssue]) -> None:
    if headers != DACOCE_HEADERS:
        _add_issue(
            issues,
            "ERROR",
            1,
            None,
            "Los encabezados de la plantilla no coinciden con la estructura DACOCE esperada.",
            f"Esperado={DACOCE_HEADERS}; Encontrado={headers}",
        )


def _validate_fixed_fields(
    row: dict[str, object],
    row_number: int,
    period_year: str,
    period_month: str,
    seen_keys: set[tuple[str, str]],
    issues: list[DacoceTemplateIssue],
) -> str | None:
    canoreg = _clean_text(row.get("CANOREG"))
    cmesreg = _clean_text(row.get("CMESREG"))
    ccodemp = _clean_upper(row.get("CCODEMP"))
    ccodcon = _clean_text(row.get("CCODCON"))
    ctipcen = _clean_upper(row.get("CTIPCEN"))

    if canoreg != period_year:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CANOREG",
            "El año de la fila no coincide con el periodo indicado.",
            canoreg,
        )

    if cmesreg != period_month:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CMESREG",
            "El mes de la fila no coincide con el periodo indicado.",
            cmesreg,
        )

    if ccodemp != "EGASA":
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODEMP",
            "La empresa debe ser EGASA.",
            ccodemp,
        )

    if not ccodcon:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODCON",
            "Código de central vacío.",
            ccodcon,
        )

    if ctipcen not in CENTER_TYPES:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CTIPCEN",
            "El tipo de central debe ser H o T.",
            ctipcen,
        )
        return None

    key = (ccodcon, ctipcen)

    if key in seen_keys:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODCON/CTIPCEN",
            "Central duplicada en la plantilla para el mismo tipo.",
            f"{ccodcon} / {ctipcen}",
        )

    seen_keys.add(key)

    return ctipcen


def _validate_numeric_fields(
    row: dict[str, object],
    row_number: int,
    issues: list[DacoceTemplateIssue],
) -> None:
    for field in NUMERIC_FIELDS:
        value = row.get(field)

        if _clean_text(value) == "":
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Campo numérico obligatorio vacío.",
                value,
            )
            continue

        decimal_value = _to_decimal(value)

        if decimal_value is None:
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Valor numérico inválido.",
                value,
            )
            continue

        if field in REQUIRED_NON_NEGATIVE_FIELDS and decimal_value < 0:
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Valor negativo no permitido.",
                decimal_value,
            )


def validate_dacoce_template(
    template_path: Path,
    period: str,
) -> DacoceTemplateValidationResult:
    period_year, period_month = parse_period(period)
    issues: list[DacoceTemplateIssue] = []

    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    workbook = load_workbook(template_path, data_only=False)

    if "DACOCE" not in workbook.sheetnames:
        raise ValueError("La plantilla no contiene una hoja llamada 'DACOCE'.")

    sheet = workbook["DACOCE"]

    headers = [
        _clean_text(sheet.cell(row=1, column=column_index).value)
        for column_index in range(1, len(DACOCE_HEADERS) + 1)
    ]
    _validate_headers(headers, issues)

    seen_keys: set[tuple[str, str]] = set()
    rows_read = 0
    hydro_rows = 0
    thermal_rows = 0

    for row_number in range(2, sheet.max_row + 1):
        row = {
            header: sheet.cell(row=row_number, column=column_index).value
            for column_index, header in enumerate(DACOCE_HEADERS, start=1)
        }

        if all(_clean_text(value) == "" for value in row.values()):
            continue

        rows_read += 1

        center_type = _validate_fixed_fields(
            row=row,
            row_number=row_number,
            period_year=period_year,
            period_month=period_month,
            seen_keys=seen_keys,
            issues=issues,
        )

        if center_type == "H":
            hydro_rows += 1
        elif center_type == "T":
            thermal_rows += 1

        _validate_numeric_fields(row=row, row_number=row_number, issues=issues)

    if rows_read == 0:
        _add_issue(
            issues,
            "ERROR",
            None,
            None,
            "La plantilla DACOCE no contiene filas de datos.",
        )

    return DacoceTemplateValidationResult(
        template_path=template_path,
        period=period,
        rows_read=rows_read,
        hydro_rows=hydro_rows,
        thermal_rows=thermal_rows,
        issues=issues,
    )


def dacoce_template_validation_to_markdown(
    result: DacoceTemplateValidationResult,
    max_issue_rows: int = 300,
) -> str:
    lines: list[str] = []

    lines.append(f"# Validación de plantilla DACOCE: `{result.template_path.name}`")
    lines.append("")
    lines.append("## Resumen")
    lines.append("")
    lines.append("| Métrica | Valor |")
    lines.append("|---|---:|")
    lines.append(f"| Plantilla | `{result.template_path}` |")
    lines.append(f"| Periodo | {result.period} |")
    lines.append(f"| Filas leídas | {result.rows_read} |")
    lines.append(f"| Filas hidroeléctricas | {result.hydro_rows} |")
    lines.append(f"| Filas termoeléctricas | {result.thermal_rows} |")
    lines.append(f"| Errores | {len(result.errors)} |")
    lines.append(f"| Advertencias | {len(result.warnings)} |")
    lines.append("")
    lines.append("## Hallazgos")
    lines.append("")

    if not result.issues:
        lines.append("- No se detectaron errores ni advertencias.")
        lines.append("")
        return "\n".join(lines)

    lines.append("| Severidad | Fila | Campo | Mensaje | Valor |")
    lines.append("|---|---:|---|---|---|")

    for issue in result.issues[:max_issue_rows]:
        row_number = "" if issue.row_number is None else str(issue.row_number)
        field = "" if issue.field is None else issue.field
        value = "" if issue.value is None else issue.value
        lines.append(
            f"| {issue.severity} | {row_number} | {field} | {issue.message} | {value} |"
        )

    if len(result.issues) > max_issue_rows:
        lines.append("")
        lines.append(
            f"> El reporte muestra los primeros {max_issue_rows} hallazgos de "
            f"{len(result.issues)} encontrados."
        )

    lines.append("")

    return "\n".join(lines)


def write_dacoce_template_validation_markdown(
    result: DacoceTemplateValidationResult,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        dacoce_template_validation_to_markdown(result),
        encoding="utf-8",
    )