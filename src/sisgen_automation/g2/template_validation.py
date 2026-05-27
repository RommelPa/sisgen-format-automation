from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.g2.catalog import load_g2_catalog
from sisgen_automation.g2.template import VEPOEN_HEADERS, parse_period


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class VepoenTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class VepoenTemplateRecord:
    canoreg: str
    cmesreg: str
    ccodemp: str
    ccoddis: str
    csieldi: str
    ccosues: str
    cnombar: str
    cclreli: str
    ntenbar: Decimal
    nprpoba: Decimal
    nprenho: Decimal
    nprenfu: Decimal
    nenveho: Decimal
    nenvefu: Decimal
    nenveto: Decimal
    npocoho: Decimal
    npocofu: Decimal
    nmadeho: Decimal
    nmadefu: Decimal
    nfacpot: Decimal
    nfacene: Decimal


@dataclass(frozen=True)
class VepoenTemplateValidationResult:
    template_path: Path
    period: str
    records: list[VepoenTemplateRecord]
    issues: list[VepoenTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _decimal_or_zero(value: Any) -> Decimal:
    parsed = _parse_decimal(value)

    if parsed is None:
        return Decimal("0")

    return parsed


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected = [header.upper() for header in VEPOEN_HEADERS]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(VEPOEN_HEADERS)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados VEPOEN en la plantilla.")


def _add_issue(
    issues: list[VepoenTemplateIssue],
    *,
    severity: IssueSeverity,
    row: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        VepoenTemplateIssue(
            severity=severity,
            row=row,
            field=field,
            message=message,
            value=value,
        )
    )


def _require_decimal(
    *,
    raw: dict[str, Any],
    field: str,
    row_index: int,
    issues: list[VepoenTemplateIssue],
) -> Decimal:
    parsed = _parse_decimal(raw[field])

    if parsed is None:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="Valor numérico obligatorio vacío o inválido.",
            value=raw[field],
        )
        return Decimal("0")

    if parsed < Decimal("0"):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="El valor numérico no puede ser negativo.",
            value=parsed,
        )

    return parsed


def validate_vepoen_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> VepoenTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla VEPOEN: {template_path}")

    expected_year, expected_month = parse_period(period)
    catalog = load_g2_catalog(catalog_path)

    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows)

    issues: list[VepoenTemplateIssue] = []
    records: list[VepoenTemplateRecord] = []
    seen_keys: set[tuple[str, str, str, str, str, str, str]] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(VEPOEN_HEADERS)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(VEPOEN_HEADERS, values, strict=True))

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccodemp = _normalize_text(raw["CCODEMP"])
        ccoddis = _normalize_text(raw["CCODDIS"])
        csieldi = _normalize_text(raw["CSIELDI"])
        ccosues = _normalize_text(raw["CCOSUES"])
        cnombar = _normalize_text(raw["CNOMBAR"])
        cclreli = _normalize_text(raw["CCLRELI"]).upper()

        row_has_error = False

        if canoreg != expected_year:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CANOREG",
                message="CANOREG no coincide con el periodo.",
                value=canoreg,
            )

        if cmesreg != expected_month:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CMESREG",
                message="CMESREG no coincide con el periodo.",
                value=cmesreg,
            )

        if ccodemp != catalog.company.ccodeemp:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODEMP",
                message="CCODEMP no coincide con el catálogo G2.",
                value=ccodemp,
            )

        if ccoddis not in catalog.distributors:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODDIS",
                message="CCODDIS no existe en el catálogo G2.",
                value=ccoddis,
            )

        if csieldi:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CSIELDI",
                message="CSIELDI debe permanecer vacío para G2.",
                value=csieldi,
            )

        if ccosues:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCOSUES",
                message="CCOSUES debe permanecer vacío para G2.",
                value=ccosues,
            )

        if not cnombar:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CNOMBAR",
                message="CNOMBAR es obligatorio.",
                value=cnombar,
            )

        if cclreli not in {"R", "L"}:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCLRELI",
                message="CCLRELI debe ser R o L.",
                value=cclreli,
            )

        ntenbar = _require_decimal(raw=raw, field="NTENBAR", row_index=row_index, issues=issues)
        nprpoba = _require_decimal(raw=raw, field="NPRPOBA", row_index=row_index, issues=issues)
        nprenho = _require_decimal(raw=raw, field="NPRENHO", row_index=row_index, issues=issues)
        nprenfu = _require_decimal(raw=raw, field="NPRENFU", row_index=row_index, issues=issues)
        nenveho = _require_decimal(raw=raw, field="NENVEHO", row_index=row_index, issues=issues)
        nenvefu = _require_decimal(raw=raw, field="NENVEFU", row_index=row_index, issues=issues)
        nenveto = _require_decimal(raw=raw, field="NENVETO", row_index=row_index, issues=issues)
        npocoho = _require_decimal(raw=raw, field="NPOCOHO", row_index=row_index, issues=issues)
        npocofu = _require_decimal(raw=raw, field="NPOCOFU", row_index=row_index, issues=issues)
        nmadeho = _decimal_or_zero(raw["NMADEHO"])
        nmadefu = _decimal_or_zero(raw["NMADEFU"])
        nfacpot = _require_decimal(raw=raw, field="NFACPOT", row_index=row_index, issues=issues)
        nfacene = _require_decimal(raw=raw, field="NFACENE", row_index=row_index, issues=issues)

        if nmadeho < Decimal("0"):
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="NMADEHO",
                message="NMADEHO no puede ser negativo.",
                value=nmadeho,
            )

        if nmadefu < Decimal("0"):
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="NMADEFU",
                message="NMADEFU no puede ser negativo.",
                value=nmadefu,
            )

        energy_total_expected = nenveho + nenvefu

        if abs(nenveto - energy_total_expected) > Decimal("0.001"):
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="NENVETO",
                message="NENVETO debe ser igual a NENVEHO + NENVEFU.",
                value=f"NENVETO={nenveto}; esperado={energy_total_expected}",
            )

        unique_key = (
            ccoddis,
            csieldi,
            ccosues,
            cnombar.upper(),
            cclreli,
            str(ntenbar),
            cmesreg,
        )

        if unique_key in seen_keys:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CLAVE",
                message="Fila duplicada para la misma distribuidora, barra, tipo y tensión.",
                value=" / ".join(unique_key),
            )

        seen_keys.add(unique_key)

        if not row_has_error:
            records.append(
                VepoenTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccodemp=ccodemp,
                    ccoddis=ccoddis,
                    csieldi=csieldi,
                    ccosues=ccosues,
                    cnombar=cnombar,
                    cclreli=cclreli,
                    ntenbar=ntenbar,
                    nprpoba=nprpoba,
                    nprenho=nprenho,
                    nprenfu=nprenfu,
                    nenveho=nenveho,
                    nenvefu=nenvefu,
                    nenveto=nenveto,
                    npocoho=npocoho,
                    npocofu=npocofu,
                    nmadeho=nmadeho,
                    nmadefu=nmadefu,
                    nfacpot=nfacpot,
                    nfacene=nfacene,
                )
            )

    if not records:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=None,
            field=None,
            message="La plantilla VEPOEN no contiene registros válidos.",
        )

    return VepoenTemplateValidationResult(
        template_path=template_path,
        period=period,
        records=records,
        issues=issues,
    )