from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from dbfread import DBF  # type: ignore[import-untyped]
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from sisgen_automation.g2.catalog import load_g2_catalog


VEPOEN_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODEMP",
    "CCODDIS",
    "CSIELDI",
    "CCOSUES",
    "CNOMBAR",
    "CCLRELI",
    "NTENBAR",
    "NPRPOBA",
    "NPRENHO",
    "NPRENFU",
    "NENVEHO",
    "NENVEFU",
    "NENVETO",
    "NPOCOHO",
    "NPOCOFU",
    "NMADEHO",
    "NMADEFU",
    "NFACPOT",
    "NFACENE",
]

EDITABLE_HEADERS = {
    "NPRPOBA",
    "NPRENHO",
    "NPRENFU",
    "NENVEHO",
    "NENVEFU",
    "NENVETO",
    "NPOCOHO",
    "NPOCOFU",
    "NMADEHO",
    "NMADEFU",
    "NFACPOT",
    "NFACENE",
}

NUMERIC_HEADERS = {
    "NTENBAR",
    "NPRPOBA",
    "NPRENHO",
    "NPRENFU",
    "NENVEHO",
    "NENVEFU",
    "NENVETO",
    "NPOCOHO",
    "NPOCOFU",
    "NMADEHO",
    "NMADEFU",
    "NFACPOT",
    "NFACENE",
}

DBF_ENCODING = "cp850"


@dataclass(frozen=True)
class VepoenTemplateResult:
    period: str
    base_period: str
    output_path: Path
    rows: int


def parse_period(period: str) -> tuple[str, str]:
    try:
        year_text, month_text = period.split("-", maxsplit=1)
    except ValueError as exc:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.") from exc

    if len(year_text) != 4 or len(month_text) != 2:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.")

    year = int(year_text)
    month = int(month_text)

    if not 1 <= month <= 12:
        raise ValueError("El mes del periodo debe estar entre 01 y 12.")

    return f"{year % 100:02d}", f"{month:02d}"


def default_vepoen_template_path(period: str) -> Path:
    year_text, month_text = period.split("-", maxsplit=1)
    return Path("reports") / "templates" / f"VEPOEN_{year_text}_{month_text}_template.xlsx"


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: object) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0")

    try:
        return Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"Valor numérico inválido en VEPOEN: {value}") from exc


def _read_vepoen_rows(source_dbf_path: Path) -> list[dict[str, object]]:
    if not source_dbf_path.exists():
        raise ValueError(f"No existe VEPOEN.DBF fuente: {source_dbf_path}")

    table = DBF(
        str(source_dbf_path),
        load=True,
        encoding=DBF_ENCODING,
        char_decode_errors="ignore",
    )

    rows: list[dict[str, object]] = []

    for record in table:
        rows.append({header: record.get(header) for header in VEPOEN_HEADERS})

    return rows


def _latest_period(rows: list[dict[str, object]]) -> str:
    periods: set[tuple[int, int]] = set()

    for row in rows:
        year = _clean_text(row.get("CANOREG"))
        month = _clean_text(row.get("CMESREG")).zfill(2)

        if not year.isdigit() or not month.isdigit():
            continue

        month_int = int(month)

        if not 1 <= month_int <= 12:
            continue

        periods.add((2000 + int(year), month_int))

    if not periods:
        raise ValueError("No se encontraron periodos válidos en VEPOEN.DBF.")

    year, month = max(periods)
    return f"{year:04d}-{month:02d}"


def _rows_for_period(rows: list[dict[str, object]], period: str) -> list[dict[str, object]]:
    year_short, month = parse_period(period)

    result: list[dict[str, object]] = []

    for row in rows:
        if _clean_text(row.get("CANOREG")) != year_short:
            continue

        if _clean_text(row.get("CMESREG")).zfill(2) != month:
            continue

        result.append(row)

    return result


def create_vepoen_template(
    *,
    period: str,
    source_dbf_path: Path,
    catalog_path: Path,
    base_period: str | None = None,
    output_path: Path | None = None,
) -> VepoenTemplateResult:
    year_short, month = parse_period(period)
    catalog = load_g2_catalog(catalog_path)

    rows = _read_vepoen_rows(source_dbf_path)

    if base_period is None:
        base_period = _latest_period(rows)

    base_rows = _rows_for_period(rows, base_period)

    if not base_rows:
        raise ValueError(f"No se encontraron registros VEPOEN para el periodo base {base_period}.")

    if output_path is None:
        output_path = default_vepoen_template_path(period)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VEPOEN"

    header_fill = PatternFill("solid", fgColor="7030A0")
    locked_fill = PatternFill("solid", fgColor="DDEBF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 1

    for col_index, header in enumerate(VEPOEN_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

    for row_index, base_row in enumerate(base_rows, start=header_row + 1):
        ccoddis = _clean_text(base_row.get("CCODDIS"))

        if ccoddis not in catalog.distributors:
            raise ValueError(
                f"VEPOEN base contiene CCODDIS no registrado en catálogo G2: {ccoddis}"
            )

        values: dict[str, Any] = {}

        for header in VEPOEN_HEADERS:
            value = base_row.get(header)

            if header == "CANOREG":
                values[header] = year_short
            elif header == "CMESREG":
                values[header] = month
            elif header in NUMERIC_HEADERS:
                values[header] = _parse_decimal(value)
            else:
                values[header] = _clean_text(value)

        for col_index, header in enumerate(VEPOEN_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=values[header])
            cell.border = border

            if header in EDITABLE_HEADERS:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

            if header in NUMERIC_HEADERS:
                cell.number_format = "0.00000"

    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 12,
        "G": 32,
        "H": 10,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 14,
        "Q": 14,
        "R": 14,
        "S": 14,
        "T": 14,
        "U": 14,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    for col_index in range(1, len(VEPOEN_HEADERS) + 1):
        sheet.cell(row=header_row, column=col_index).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:U{header_row + len(base_rows)}"
    sheet.protection.sheet = True
    sheet.protection.enable()

    workbook.save(output_path)

    return VepoenTemplateResult(
        period=period,
        base_period=base_period,
        output_path=output_path,
        rows=len(base_rows),
    )