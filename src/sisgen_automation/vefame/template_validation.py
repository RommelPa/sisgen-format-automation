from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from enum import Enum
from pathlib import Path
from openpyxl import load_workbook

from sisgen_automation.g8.catalog import load_g8_catalog
from sisgen_automation.vefame.template import (
    CALCULATED_HEADERS,
    EDITABLE_HEADERS,
    NUMERIC_HEADERS,
    VEFAME_HEADERS,
    parse_period,
)


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class VefameTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class VefameTemplateRecord:
    row: int
    values: dict[str, str | Decimal]


@dataclass(frozen=True)
class VefameTemplateValidationResult:
    period: str
    records: tuple[VefameTemplateRecord, ...]
    issues: tuple[VefameTemplateIssue, ...]

    @property
    def errors(self) -> tuple[VefameTemplateIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warnings(self) -> tuple[VefameTemplateIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _normalize_year(value: object) -> str:
    text = _clean_text(value)

    if text.endswith(".0"):
        text = text[:-2]

    return text.zfill(2)


def _normalize_month(value: object) -> str:
    text = _clean_text(value)

    if text.endswith(".0"):
        text = text[:-2]

    return text.zfill(2)


def _decimal_or_none(value: object) -> Decimal | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if text.startswith("="):
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _required_decimal(
    *,
    row: int,
    field: str,
    value: object,
    issues: list[VefameTemplateIssue],
) -> Decimal:
    parsed = _decimal_or_none(value)

    if parsed is None:
        issues.append(
            VefameTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=row,
                field=field,
                message="Valor numérico obligatorio vacío o inválido.",
                value=value,
            )
        )
        return Decimal("0")

    return parsed


def _optional_decimal(value: object) -> Decimal:
    parsed = _decimal_or_none(value)
    if parsed is None:
        return Decimal("0")
    return parsed


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)


def _calculated_values(values: dict[str, str | Decimal]) -> dict[str, Decimal]:
    nenacho = Decimal(values["NENACHO"])
    nenacfu = Decimal(values["NENACFU"])
    nhrspun = Decimal(values["NHRSPUN"])
    nfuhopu = Decimal(values["NFUHOPU"])
    nexhopu = Decimal(values["NEXHOPU"])
    nexfuho = Decimal(values["NEXFUHO"])
    nprpoho = Decimal(values["NPRPOHO"])
    nprpofu = Decimal(values["NPRPOFU"])
    nprpoex = Decimal(values["NPRPOEX"])

    return {
        "NENACTO": _money(nenacho + nenacfu),
        "NFACPOT": _money((nhrspun * Decimal("1000") * nprpoho) + (nfuhopu * Decimal("1000") * nprpofu)),
        "NFACEXC": _money((nexhopu * Decimal("1000") * nprpoex) + (nexfuho * Decimal("1000") * nprpoex)),
    }


def _validate_headers(headers: list[str], issues: list[VefameTemplateIssue]) -> None:
    if headers != VEFAME_HEADERS:
        issues.append(
            VefameTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=1,
                field=None,
                message="La plantilla VEFAME no tiene las columnas esperadas o están en otro orden.",
                value=", ".join(headers),
            )
        )


def validate_vefame_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> VefameTemplateValidationResult:
    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla VEFAME: {template_path}")

    year_short, month = parse_period(period)
    catalog = load_g8_catalog(catalog_path)
    clients_by_code = catalog.clients_by_code

    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook.active

    issues: list[VefameTemplateIssue] = []
    headers = [_clean_text(sheet.cell(row=1, column=col).value) for col in range(1, len(VEFAME_HEADERS) + 1)]
    _validate_headers(headers, issues)

    if issues:
        return VefameTemplateValidationResult(period=period, records=(), issues=tuple(issues))

    records: list[VefameTemplateRecord] = []
    seen_clients: set[str] = set()

    for row_index in range(2, sheet.max_row + 1):
        raw_values = {
            header: sheet.cell(row=row_index, column=col_index).value
            for col_index, header in enumerate(VEFAME_HEADERS, start=1)
        }

        if all(value is None or str(value).strip() == "" for value in raw_values.values()):
            continue

        values: dict[str, str | Decimal] = {}

        canoreg = _normalize_year(raw_values["CANOREG"])
        cmesreg = _normalize_month(raw_values["CMESREG"])
        ccodeemp = _clean_text(raw_values["CCODEMP"])
        cnivten = _clean_text(raw_values["CNIVTEN"])
        ccoclli = _clean_text(raw_values["CCOCLLI"])
        cpotmax = _clean_text(raw_values["CPOTMAX"])

        if canoreg != year_short:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CANOREG",
                    message=f"Año inválido. Se esperaba {year_short}.",
                    value=raw_values["CANOREG"],
                )
            )

        if cmesreg != month:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CMESREG",
                    message=f"Mes inválido. Se esperaba {month}.",
                    value=raw_values["CMESREG"],
                )
            )

        if ccodeemp != catalog.company.ccodeemp:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODEMP",
                    message=f"Empresa inválida. Se esperaba {catalog.company.ccodeemp}.",
                    value=ccodeemp,
                )
            )

        client = clients_by_code.get(ccoclli)
        if client is None:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCOCLLI",
                    message="Cliente libre no existe en catálogo G8.",
                    value=ccoclli,
                )
            )
        else:
            if ccoclli in seen_clients:
                issues.append(
                    VefameTemplateIssue(
                        severity=IssueSeverity.ERROR,
                        row=row_index,
                        field="CCOCLLI",
                        message="Cliente libre duplicado en plantilla.",
                        value=ccoclli,
                    )
                )

            seen_clients.add(ccoclli)

            if cnivten != client.cnivten:
                issues.append(
                    VefameTemplateIssue(
                        severity=IssueSeverity.ERROR,
                        row=row_index,
                        field="CNIVTEN",
                        message=f"Nivel de tensión inválido. Se esperaba {client.cnivten}.",
                        value=cnivten,
                    )
                )

        if not cpotmax:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CPOTMAX",
                    message="CPOTMAX es obligatorio.",
                    value=cpotmax,
                )
            )

        values["CANOREG"] = year_short
        values["CMESREG"] = month
        values["CCODEMP"] = ccodeemp
        values["CNIVTEN"] = cnivten
        values["CCOCLLI"] = ccoclli
        values["CPOTMAX"] = cpotmax

        for header in VEFAME_HEADERS:
            if header in values:
                continue

            raw_value = raw_values[header]

            if header in EDITABLE_HEADERS:
                values[header] = _required_decimal(
                    row=row_index,
                    field=header,
                    value=raw_value,
                    issues=issues,
                )
            elif header in CALCULATED_HEADERS:
                values[header] = _optional_decimal(raw_value)
            elif header in NUMERIC_HEADERS:
                values[header] = _required_decimal(
                    row=row_index,
                    field=header,
                    value=raw_value,
                    issues=issues,
                )
            else:
                values[header] = _clean_text(raw_value)

        calculated = _calculated_values(values)

        for field, expected in calculated.items():
            raw_value = raw_values[field]
            current = _decimal_or_none(raw_value)

            if current is not None and abs(current - expected) > Decimal("0.00001"):
                issues.append(
                    VefameTemplateIssue(
                        severity=IssueSeverity.WARNING,
                        row=row_index,
                        field=field,
                        message="Valor calculado no coincide con los campos base. Se usará el cálculo interno.",
                        value=f"{current} != {expected}",
                    )
                )

            values[field] = expected

        records.append(VefameTemplateRecord(row=row_index, values=values))

    for client in catalog.free_clients:
        if client.ccoclli not in seen_clients:
            issues.append(
                VefameTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=None,
                    field="CCOCLLI",
                    message="Cliente del catálogo no aparece en la plantilla VEFAME.",
                    value=f"{client.ccoclli} - {client.name}",
                )
            )

    valid_records = tuple(records) if not any(issue.severity == IssueSeverity.ERROR for issue in issues) else ()

    return VefameTemplateValidationResult(
        period=period,
        records=valid_records,
        issues=tuple(issues),
    )
