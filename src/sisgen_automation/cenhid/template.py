from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from sisgen_automation.cenhid.catalog import CenhidUnit, load_cenhid_catalog

CENHID_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODEMP",
    "CCODCON",
    "CCODCEN",
    "CTIPGRU",
    "CNOMNUM",
    "CESTGRU",
    "NPOTINS",
    "NPOTEFE",
    "NHORPUN",
    "NFUHOPU",
    "NTOPRBR",
    "CHRSMAN",
    "CHRSOPE",
    "CHRSSAL",
    "CENTRAL",
    "GRUPO",
]

EDITABLE_COLUMNS = {"H", "K", "L", "N", "O", "P"}


def parse_period(period: str) -> tuple[str, str]:
    if not re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])", period):
        raise ValueError("El periodo debe tener formato YYYY-MM. Ejemplo válido: 2026-01")

    year = period[2:4]
    month = period[5:7]

    return year, month


def _decimal_to_float(value: Decimal) -> float:
    return float(value)


def _write_instructions_sheet(workbook: Workbook, period: str) -> None:
    sheet = workbook.create_sheet("INSTRUCCIONES")

    rows = [
        ("Plantilla", "CENHID"),
        ("Periodo", period),
        ("Uso", "Completar solo las columnas editables."),
        ("Columnas editables", "CESTGRU, NHORPUN, NFUHOPU, CHRSMAN, CHRSOPE, CHRSSAL."),
        ("Columna calculada", "NTOPRBR se calcula automáticamente como NHORPUN + NFUHOPU."),
        ("Advertencia", "No modificar códigos, empresa, tipo de grupo ni nombres de grupo."),
        ("Estado CESTGRU", "Usar S para servicio o N para no servicio."),
        ("Unidades energía", "MWh."),
        ("Unidades horas", "h."),
    ]

    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 90

    for cell in sheet["A"]:
        cell.font = Font(bold=True)


def _write_header(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(CENHID_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(CENHID_HEADERS))}1"


def _write_unit_row(
    sheet,
    row_number: int,
    year: str,
    month: str,
    unit: CenhidUnit,
) -> None:
    values = [
        year,
        month,
        "EGASA",
        unit.ccodcon,
        unit.ccodcen,
        "HI",
        unit.cnomnum,
        "S",
        _decimal_to_float(unit.npotins),
        _decimal_to_float(unit.npotefe),
        None,
        None,
        f"=K{row_number}+L{row_number}",
        None,
        None,
        None,
        unit.central,
        unit.group,
    ]

    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_number, column=column_index, value=value)

        column_letter = get_column_letter(column_index)

        if column_letter in EDITABLE_COLUMNS:
            cell.protection = Protection(locked=False)
        else:
            cell.protection = Protection(locked=True)

        if column_letter in {"I", "J", "K", "L", "M", "N", "O", "P"}:
            cell.number_format = "0.00000"


def _apply_sheet_format(sheet, last_row: int) -> None:
    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 10,
        "G": 14,
        "H": 10,
        "I": 14,
        "J": 14,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 18,
        "R": 14,
    }

    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    fixed_fill = PatternFill("solid", fgColor="D9EAF7")
    formula_fill = PatternFill("solid", fgColor="E2F0D9")

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    for row in sheet.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            column_letter = get_column_letter(cell.column)

            if column_letter in EDITABLE_COLUMNS:
                cell.fill = editable_fill
            elif column_letter == "M":
                cell.fill = formula_fill
            else:
                cell.fill = fixed_fill

            cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet["H1"].comment = Comment("Estado del grupo: S o N.", "SISGEN")
    sheet["M1"].comment = Comment("Campo calculado: NHORPUN + NFUHOPU.", "SISGEN")

    status_validation = DataValidation(
        type="list",
        formula1='"S,N"',
        allow_blank=False,
    )
    status_validation.error = "Solo se permite S o N."
    status_validation.errorTitle = "Estado inválido"
    status_validation.prompt = "Seleccione S o N."
    status_validation.promptTitle = "CESTGRU"

    sheet.add_data_validation(status_validation)
    status_validation.add(f"H2:H{last_row}")

    sheet.protection.sheet = True
    sheet.protection.password = "sisgen"


def create_cenhid_template(
    period: str,
    catalog_path: Path,
    output_path: Path | None = None,
) -> Path:
    year, month = parse_period(period)
    catalog = load_cenhid_catalog(catalog_path)

    year_text, month_text = period.split("-", maxsplit=1)
    if output_path is None:
        output_path = Path("reports") / "templates" / f"CENHID_{year_text}_{month_text}_template.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CENHID"

    _write_header(sheet)

    for row_number, unit in enumerate(catalog.values(), start=2):
        _write_unit_row(sheet, row_number=row_number, year=year, month=month, unit=unit)

    last_row = len(catalog) + 1
    _apply_sheet_format(sheet, last_row=last_row)
    _write_instructions_sheet(workbook, period=period)

    workbook.save(output_path)

    return output_path