from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Mapping

from dbfread import DBF  # type: ignore[import-untyped]
from openpyxl import load_workbook

from sisgen_automation.cenhid.template import CENHID_HEADERS, parse_period
from sisgen_automation.cenhid.template_validation import validate_cenhid_template
from sisgen_automation.dbf.profile import DbfField, read_dbf_profile
from sisgen_automation.dbf.compatibility import assert_sisgen_expected_layout

DBF_ENCODING = "cp850"
DBF_EOF_MARKER = b"\x1A"


@dataclass(frozen=True)
class CenhidExportResult:
    source_dbf_path: Path
    template_path: Path
    output_path: Path
    period: str
    original_record_count: int
    appended_record_count: int
    final_record_count: int


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _to_decimal(value: object) -> Decimal:
    if value is None:
        raise ValueError("Valor numérico vacío.")

    text = str(value).strip()

    if not text:
        raise ValueError("Valor numérico vacío.")

    return Decimal(text)


def _period_exists_in_dbf(source_dbf_path: Path, period: str) -> bool:
    year, month = parse_period(period)
    table = DBF(str(source_dbf_path), load=True, char_decode_errors="ignore")

    for record in table:
        if _clean_text(record.get("CANOREG")) == year and _clean_text(record.get("CMESREG")) == month:
            return True

    return False


def _read_template_rows(template_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook["CENHID"]

    rows: list[dict[str, object]] = []

    for row_number in range(2, sheet.max_row + 1):
        row = {
            header: sheet.cell(row=row_number, column=column_index).value
            for column_index, header in enumerate(CENHID_HEADERS, start=1)
        }

        if all(_clean_text(value) == "" for value in row.values()):
            continue

        nhorpun = _to_decimal(row["NHORPUN"])
        nfuhopu = _to_decimal(row["NFUHOPU"])

        row["NTOPRBR"] = nhorpun + nfuhopu

        rows.append(row)

    return rows


def _serialize_character(value: object, field: DbfField) -> bytes:
    text = _clean_text(value)

    encoded = text.encode(DBF_ENCODING, errors="replace")

    if len(encoded) > field.length:
        raise ValueError(
            f"El valor '{text}' excede la longitud del campo {field.name}: "
            f"{len(encoded)} > {field.length}"
        )

    return encoded.ljust(field.length, b" ")


def _serialize_numeric(value: object, field: DbfField) -> bytes:
    decimal_value = _to_decimal(value)

    quantizer = Decimal("1").scaleb(-field.decimal_count)
    rounded_value = decimal_value.quantize(quantizer, rounding=ROUND_HALF_UP)

    text = f"{rounded_value:.{field.decimal_count}f}"

    if len(text) > field.length:
        raise ValueError(
            f"El valor numérico '{text}' excede la longitud del campo {field.name}: "
            f"{len(text)} > {field.length}"
        )

    return text.rjust(field.length, " ").encode("ascii")


def _serialize_field(value: object, field: DbfField) -> bytes:
    if field.field_type == "C":
        return _serialize_character(value, field)

    if field.field_type == "N":
        return _serialize_numeric(value, field)

    raise ValueError(f"Tipo de campo DBF no soportado para exportación: {field.field_type}")


def _serialize_record(row: Mapping[str, object], fields: list[DbfField]) -> bytes:
    record = bytearray()
    record.extend(b" ")

    for field in fields:
        record.extend(_serialize_field(row.get(field.name), field))

    return bytes(record)


def _write_appended_dbf(
    source_dbf_path: Path,
    output_path: Path,
    new_rows: list[dict[str, object]],
) -> CenhidExportResult:
    profile = read_dbf_profile(source_dbf_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    source_bytes = source_dbf_path.read_bytes()

    expected_data_end = profile.header_length + (profile.record_count * profile.record_length)

    if len(source_bytes) < expected_data_end:
        raise ValueError(
            "El DBF fuente es más pequeño que lo indicado por su cabecera. "
            "No es seguro exportar."
        )

    header = bytearray(source_bytes[: profile.header_length])
    existing_records = source_bytes[profile.header_length:expected_data_end]

    serialized_records = [_serialize_record(row, profile.fields) for row in new_rows]

    for index, record in enumerate(serialized_records, start=1):
        if len(record) != profile.record_length:
            raise ValueError(
                f"El registro nuevo #{index} tiene longitud inválida: "
                f"{len(record)} != {profile.record_length}"
            )
        
    source_has_eof = (
        len(source_bytes) > expected_data_end
        and source_bytes[expected_data_end : expected_data_end + 1] == DBF_EOF_MARKER
    )

    final_record_count = profile.record_count + len(serialized_records)

    # Compatibilidad legacy:
    # conservamos la cabecera original y solo actualizamos el contador de registros.
    header[4:8] = final_record_count.to_bytes(4, byteorder="little")

    output_bytes = bytes(header) + existing_records + b"".join(serialized_records)

    if source_has_eof:
        output_bytes += DBF_EOF_MARKER

    output_path.write_bytes(output_bytes)

    return CenhidExportResult(
        source_dbf_path=source_dbf_path,
        template_path=Path(),
        output_path=output_path,
        period="",
        original_record_count=profile.record_count,
        appended_record_count=len(serialized_records),
        final_record_count=final_record_count,
    )


def export_cenhid_dbf(
    source_dbf_path: Path,
    template_path: Path,
    period: str,
    catalog_path: Path,
    output_path: Path | None = None,
    allow_existing_period: bool = False,
) -> CenhidExportResult:
    assert_sisgen_expected_layout(source_dbf_path, "CENHID")
    validation_result = validate_cenhid_template(
        template_path=template_path,
        period=period,
        catalog_path=catalog_path,
    )

    if validation_result.has_errors:
        raise ValueError(
            "La plantilla tiene errores. Corrige la plantilla antes de exportar el DBF."
        )

    if _period_exists_in_dbf(source_dbf_path, period) and not allow_existing_period:
        raise ValueError(
            f"El periodo {period} ya existe en el DBF histórico. "
            "No se exportó para evitar duplicar información."
        )

    new_rows = _read_template_rows(template_path)

    if output_path is None:
        output_path = Path("data") / "output" / period / "CENHID.DBF"

    result = _write_appended_dbf(
        source_dbf_path=source_dbf_path,
        output_path=output_path,
        new_rows=new_rows,
    )

    return CenhidExportResult(
        source_dbf_path=source_dbf_path,
        template_path=template_path,
        output_path=result.output_path,
        period=period,
        original_record_count=result.original_record_count,
        appended_record_count=result.appended_record_count,
        final_record_count=result.final_record_count,
    )