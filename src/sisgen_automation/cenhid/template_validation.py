from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from sisgen_automation.cenhid.catalog import CenhidUnit, load_cenhid_catalog
from sisgen_automation.cenhid.template import CENHID_HEADERS, parse_period

IssueSeverity = Literal["ERROR", "WARNING"]

REQUIRED_MONTHLY_FIELDS = [
    "NHORPUN",
    "NFUHOPU",
    "CHRSMAN",
    "CHRSOPE",
    "CHRSSAL",
]

NUMERIC_FIELDS = [
    "NPOTINS",
    "NPOTEFE",
    "NHORPUN",
    "NFUHOPU",
    "CHRSMAN",
    "CHRSOPE",
    "CHRSSAL",
]

STATUS_VALUES = {"S", "N"}


@dataclass(frozen=True)
class TemplateIssue:
    severity: IssueSeverity
    row_number: int | None
    field: str | None
    message: str
    value: str | None = None


@dataclass(frozen=True)
class CenhidTemplateValidationResult:
    template_path: Path
    period: str
    catalog_path: Path
    rows_read: int
    expected_units: int
    valid_units: int
    issues: list[TemplateIssue]

    @property
    def errors(self) -> list[TemplateIssue]:
        return [issue for issue in self.issues if issue.severity == "ERROR"]

    @property
    def warnings(self) -> list[TemplateIssue]:
        return [issue for issue in self.issues if issue.severity == "WARNING"]

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


def _add_issue(
    issues: list[TemplateIssue],
    severity: IssueSeverity,
    row_number: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        TemplateIssue(
            severity=severity,
            row_number=row_number,
            field=field,
            message=message,
            value=None if value is None else str(value),
        )
    )


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _clean_upper(value: object) -> str:
    return _clean_text(value).upper()


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _validate_headers(headers: list[str], issues: list[TemplateIssue]) -> None:
    if headers != CENHID_HEADERS:
        _add_issue(
            issues,
            "ERROR",
            1,
            None,
            "Los encabezados de la plantilla no coinciden con la estructura esperada.",
            f"Esperado={CENHID_HEADERS}; Encontrado={headers}",
        )


def _validate_fixed_fields(
    row: dict[str, object],
    row_number: int,
    period_year: str,
    period_month: str,
    catalog: dict[tuple[str, str], CenhidUnit],
    seen_units: set[tuple[str, str]],
    issues: list[TemplateIssue],
) -> tuple[str, str] | None:
    canoreg = _clean_text(row.get("CANOREG"))
    cmesreg = _clean_text(row.get("CMESREG"))
    ccodemp = _clean_upper(row.get("CCODEMP"))
    ccodcon = _clean_text(row.get("CCODCON"))
    ccodcen = _clean_text(row.get("CCODCEN"))
    ctipgru = _clean_upper(row.get("CTIPGRU"))
    cnomnum = _clean_text(row.get("CNOMNUM"))

    if canoreg != period_year:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CANOREG",
            "El año de la fila no coincide con el periodo indicado.",
            canoreg,
        )

    if cmesreg != period_month:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CMESREG",
            "El mes de la fila no coincide con el periodo indicado.",
            cmesreg,
        )

    if ccodemp != "EGASA":
        _add_issue(issues, "ERROR", row_number, "CCODEMP", "La empresa debe ser EGASA.", ccodemp)

    if ccodcen != ccodcon:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODCEN",
            "CCODCEN debe ser igual a CCODCON.",
            f"CCODCON={ccodcon}; CCODCEN={ccodcen}",
        )

    if ctipgru != "HI":
        _add_issue(issues, "ERROR", row_number, "CTIPGRU", "El tipo de grupo debe ser HI.", ctipgru)

    key = (ccodcon, cnomnum)

    if key not in catalog:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODCON/CNOMNUM",
            "La unidad no existe en el catálogo CENHID.",
            f"{ccodcon} / {cnomnum}",
        )
        return None

    if key in seen_units:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CCODCON/CNOMNUM",
            "Unidad duplicada en la plantilla.",
            f"{ccodcon} / {cnomnum}",
        )

    seen_units.add(key)

    return key


def _validate_status(
    row: dict[str, object],
    row_number: int,
    issues: list[TemplateIssue],
) -> None:
    status = _clean_upper(row.get("CESTGRU"))

    if status not in STATUS_VALUES:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "CESTGRU",
            "El estado del grupo debe ser S o N.",
            row.get("CESTGRU"),
        )


def _validate_numeric_fields(
    row: dict[str, object],
    row_number: int,
    issues: list[TemplateIssue],
) -> None:
    for field in REQUIRED_MONTHLY_FIELDS:
        if _clean_text(row.get(field)) == "":
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Campo mensual obligatorio vacío.",
                row.get(field),
            )

    for field in NUMERIC_FIELDS:
        value = row.get(field)
        decimal_value = _to_decimal(value)

        if decimal_value is None:
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Valor numérico inválido o vacío.",
                value,
            )
            continue

        if decimal_value < 0:
            _add_issue(
                issues,
                "ERROR",
                row_number,
                field,
                "Valor negativo no permitido.",
                decimal_value,
            )


def _validate_total_formula(
    row: dict[str, object],
    row_number: int,
    issues: list[TemplateIssue],
) -> None:
    expected_formula = f"=K{row_number}+L{row_number}"
    actual_formula = row.get("NTOPRBR")

    if actual_formula != expected_formula:
        _add_issue(
            issues,
            "ERROR",
            row_number,
            "NTOPRBR",
            "La fórmula de energía total fue modificada o eliminada.",
            actual_formula,
        )


def validate_cenhid_template(
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> CenhidTemplateValidationResult:
    period_year, period_month = parse_period(period)
    catalog = load_cenhid_catalog(catalog_path)
    issues: list[TemplateIssue] = []

    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    workbook = load_workbook(template_path, data_only=False)

    if "CENHID" not in workbook.sheetnames:
        raise ValueError("La plantilla no contiene una hoja llamada 'CENHID'.")

    sheet = workbook["CENHID"]

    headers = [_clean_text(sheet.cell(row=1, column=column_index).value) for column_index in range(1, 19)]
    _validate_headers(headers, issues)

    seen_units: set[tuple[str, str]] = set()
    rows_read = 0

    for row_number in range(2, sheet.max_row + 1):
        row = {
            header: sheet.cell(row=row_number, column=column_index).value
            for column_index, header in enumerate(CENHID_HEADERS, start=1)
        }

        if all(_clean_text(value) == "" for value in row.values()):
            continue

        rows_read += 1

        unit_key = _validate_fixed_fields(
            row=row,
            row_number=row_number,
            period_year=period_year,
            period_month=period_month,
            catalog=catalog,
            seen_units=seen_units,
            issues=issues,
        )

        _validate_status(row, row_number, issues)
        _validate_numeric_fields(row, row_number, issues)
        _validate_total_formula(row, row_number, issues)

        if unit_key is not None:
            catalog_unit = catalog[unit_key]
            central = _clean_text(row.get("CENTRAL"))
            group = _clean_text(row.get("GRUPO"))

            if central != catalog_unit.central:
                _add_issue(
                    issues,
                    "WARNING",
                    row_number,
                    "CENTRAL",
                    "El nombre de central no coincide con el catálogo.",
                    f"Excel={central}; catálogo={catalog_unit.central}",
                )

            if group != catalog_unit.group:
                _add_issue(
                    issues,
                    "WARNING",
                    row_number,
                    "GRUPO",
                    "El nombre de grupo no coincide con el catálogo.",
                    f"Excel={group}; catálogo={catalog_unit.group}",
                )

    missing_units = set(catalog) - seen_units

    for ccodcon, cnomnum in sorted(missing_units):
        _add_issue(
            issues,
            "ERROR",
            None,
            "CCODCON/CNOMNUM",
            "Falta una unidad del catálogo en la plantilla.",
            f"{ccodcon} / {cnomnum}",
        )

    return CenhidTemplateValidationResult(
        template_path=template_path,
        period=period,
        catalog_path=catalog_path,
        rows_read=rows_read,
        expected_units=len(catalog),
        valid_units=len(seen_units),
        issues=issues,
    )


def template_validation_to_markdown(
    result: CenhidTemplateValidationResult,
    max_issue_rows: int = 300,
) -> str:
    lines: list[str] = []

    lines.append(f"# Validación de plantilla CENHID: `{result.template_path.name}`")
    lines.append("")
    lines.append("## Resumen")
    lines.append("")
    lines.append("| Métrica | Valor |")
    lines.append("|---|---:|")
    lines.append(f"| Plantilla | `{result.template_path}` |")
    lines.append(f"| Periodo | {result.period} |")
    lines.append(f"| Catálogo usado | `{result.catalog_path}` |")
    lines.append(f"| Filas leídas | {result.rows_read} |")
    lines.append(f"| Unidades esperadas | {result.expected_units} |")
    lines.append(f"| Unidades válidas detectadas | {result.valid_units} |")
    lines.append(f"| Errores | {len(result.errors)} |")
    lines.append(f"| Advertencias | {len(result.warnings)} |")
    lines.append("")
    lines.append("## Hallazgos")
    lines.append("")

    if not result.issues:
        lines.append("- No se detectaron errores ni advertencias.")
        lines.append("")
        return "\n".join(lines)

    lines.append("| Severidad | Fila | Campo | Mensaje | Valor |")
    lines.append("|---|---:|---|---|---|")

    for issue in result.issues[:max_issue_rows]:
        row_number = "" if issue.row_number is None else str(issue.row_number)
        field = "" if issue.field is None else issue.field
        value = "" if issue.value is None else issue.value
        lines.append(
            f"| {issue.severity} | {row_number} | {field} | {issue.message} | {value} |"
        )

    if len(result.issues) > max_issue_rows:
        lines.append("")
        lines.append(
            f"> El reporte muestra los primeros {max_issue_rows} hallazgos de "
            f"{len(result.issues)} encontrados."
        )

    lines.append("")

    return "\n".join(lines)


def write_template_validation_markdown(
    result: CenhidTemplateValidationResult,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(template_validation_to_markdown(result), encoding="utf-8")