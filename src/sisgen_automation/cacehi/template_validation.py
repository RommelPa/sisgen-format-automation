from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.cacehi.template import CACEHI_HEADERS, parse_period
from sisgen_automation.g11.catalog import load_g11_catalog


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class CacehiTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class CacehiTemplateRecord:
    canoreg: str
    cmesreg: str
    ccodgen: str
    ccodcen: str
    cnomcen: str
    cnomcue: str
    ncamedi: Decimal
    nalvout: Decimal


@dataclass(frozen=True)
class CacehiTemplateValidationResult:
    template_path: Path
    period: str
    records: list[CacehiTemplateRecord]
    issues: list[CacehiTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected = [header.upper() for header in CACEHI_HEADERS]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(CACEHI_HEADERS)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados CACEHI en la plantilla.")


def _add_issue(
    issues: list[CacehiTemplateIssue],
    *,
    severity: IssueSeverity,
    row: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        CacehiTemplateIssue(
            severity=severity,
            row=row,
            field=field,
            message=message,
            value=value,
        )
    )


def _require_decimal(
    *,
    raw: dict[str, Any],
    field: str,
    row_index: int,
    issues: list[CacehiTemplateIssue],
) -> Decimal:
    parsed = _parse_decimal(raw[field])

    if parsed is None:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="Valor numérico obligatorio vacío o inválido.",
            value=raw[field],
        )
        return Decimal("0")

    if parsed < Decimal("0"):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="El valor numérico no puede ser negativo.",
            value=parsed,
        )

    return parsed


def validate_cacehi_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> CacehiTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla CACEHI: {template_path}")

    expected_year, expected_month = parse_period(period)
    catalog = load_g11_catalog(catalog_path)

    hydro_catalog = {
        item.ccodcen: item
        for item in catalog.hydro
    }

    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows)

    issues: list[CacehiTemplateIssue] = []
    records: list[CacehiTemplateRecord] = []
    seen_keys: set[str] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(CACEHI_HEADERS)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(CACEHI_HEADERS, values, strict=True))

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccodgen = _normalize_text(raw["CCODGEN"])
        ccodcen = _normalize_text(raw["CCODCEN"])
        cnomcen = _normalize_text(raw["CNOMCEN"])
        cnomcue = _normalize_text(raw["CNOMCUE"])

        row_has_error = False

        if canoreg != expected_year:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CANOREG",
                message="CANOREG no coincide con el periodo.",
                value=canoreg,
            )

        if cmesreg != expected_month:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CMESREG",
                message="CMESREG no coincide con el periodo.",
                value=cmesreg,
            )

        if ccodgen != catalog.company.ccodgen:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODGEN",
                message="CCODGEN no coincide con el catálogo G11.",
                value=ccodgen,
            )

        unit = hydro_catalog.get(ccodcen)

        if unit is None:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODCEN",
                message="CCODCEN no existe en el catálogo G11 hidro.",
                value=ccodcen,
            )
        else:
            if cnomcen != unit.cnomcen:
                row_has_error = True
                _add_issue(
                    issues,
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CNOMCEN",
                    message="CNOMCEN no coincide con el catálogo G11.",
                    value=cnomcen,
                )

            if cnomcue != unit.cnomcue:
                row_has_error = True
                _add_issue(
                    issues,
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CNOMCUE",
                    message="CNOMCUE no coincide con el catálogo G11.",
                    value=cnomcue,
                )

        ncamedi = _require_decimal(raw=raw, field="NCAMEDI", row_index=row_index, issues=issues)
        nalvout = _require_decimal(raw=raw, field="NALVOUT", row_index=row_index, issues=issues)

        if ccodcen in seen_keys:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODCEN",
                message="Central hidro duplicada en la plantilla.",
                value=ccodcen,
            )

        seen_keys.add(ccodcen)

        if not row_has_error:
            records.append(
                CacehiTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccodgen=ccodgen,
                    ccodcen=ccodcen,
                    cnomcen=cnomcen,
                    cnomcue=cnomcue,
                    ncamedi=ncamedi,
                    nalvout=nalvout,
                )
            )

    expected_count = len(catalog.hydro)

    if len(records) != expected_count:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=None,
            field=None,
            message="La plantilla CACEHI no contiene todas las centrales hidro esperadas.",
            value=f"válidas={len(records)}; esperadas={expected_count}",
        )

    return CacehiTemplateValidationResult(
        template_path=template_path,
        period=period,
        records=records,
        issues=issues,
    )