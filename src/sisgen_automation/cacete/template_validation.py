from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.cacete.template import CACETE_HEADERS, parse_period
from sisgen_automation.g11.catalog import load_g11_catalog


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class CaceteTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class CaceteTemplateRecord:
    canoreg: str
    cmesreg: str
    ccodgen: str
    ccodcen: str
    cnomcen: str
    ctipcom: str
    cdescom: str
    nvolalm: Decimal
    nadqmes: Decimal


@dataclass(frozen=True)
class CaceteTemplateValidationResult:
    template_path: Path
    period: str
    records: list[CaceteTemplateRecord]
    issues: list[CaceteTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected = [header.upper() for header in CACETE_HEADERS]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(CACETE_HEADERS)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados CACETE en la plantilla.")


def _add_issue(
    issues: list[CaceteTemplateIssue],
    *,
    severity: IssueSeverity,
    row: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        CaceteTemplateIssue(
            severity=severity,
            row=row,
            field=field,
            message=message,
            value=value,
        )
    )


def _require_decimal(
    *,
    raw: dict[str, Any],
    field: str,
    row_index: int,
    issues: list[CaceteTemplateIssue],
) -> Decimal:
    parsed = _parse_decimal(raw[field])

    if parsed is None:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="Valor numérico obligatorio vacío o inválido.",
            value=raw[field],
        )
        return Decimal("0")

    if parsed < Decimal("0"):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="El valor numérico no puede ser negativo.",
            value=parsed,
        )

    return parsed


def validate_cacete_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> CaceteTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla CACETE: {template_path}")

    expected_year, expected_month = parse_period(period)
    catalog = load_g11_catalog(catalog_path)

    thermal_catalog = {
        (item.ccodcen, item.ctipcom): item
        for item in catalog.thermal_fuels
    }

    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows)

    issues: list[CaceteTemplateIssue] = []
    records: list[CaceteTemplateRecord] = []
    seen_keys: set[tuple[str, str]] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(CACETE_HEADERS)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(CACETE_HEADERS, values, strict=True))

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccodgen = _normalize_text(raw["CCODGEN"])
        ccodcen = _normalize_text(raw["CCODCEN"])
        cnomcen = _normalize_text(raw["CNOMCEN"])
        ctipcom = _normalize_text(raw["CTIPCOM"])
        cdescom = _normalize_text(raw["CDESCOM"])

        row_has_error = False

        if canoreg != expected_year:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CANOREG",
                message="CANOREG no coincide con el periodo.",
                value=canoreg,
            )

        if cmesreg != expected_month:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CMESREG",
                message="CMESREG no coincide con el periodo.",
                value=cmesreg,
            )

        if ccodgen != catalog.company.ccodgen:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODGEN",
                message="CCODGEN no coincide con el catálogo G11.",
                value=ccodgen,
            )

        fuel = thermal_catalog.get((ccodcen, ctipcom))

        if fuel is None:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CLAVE",
                message="CCODCEN/CTIPCOM no existe en el catálogo G11 térmico.",
                value=f"{ccodcen} / {ctipcom}",
            )
        else:
            if cnomcen != fuel.cnomcen:
                row_has_error = True
                _add_issue(
                    issues,
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CNOMCEN",
                    message="CNOMCEN no coincide con el catálogo G11.",
                    value=cnomcen,
                )

            if cdescom != fuel.cdescom:
                row_has_error = True
                _add_issue(
                    issues,
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CDESCOM",
                    message="CDESCOM no coincide con el catálogo G11.",
                    value=cdescom,
                )

        nvolalm = _require_decimal(raw=raw, field="NVOLALM", row_index=row_index, issues=issues)
        nadqmes = _require_decimal(raw=raw, field="NADQMES", row_index=row_index, issues=issues)

        key = (ccodcen, ctipcom)

        if key in seen_keys:
            row_has_error = True
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CLAVE",
                message="Central/combustible duplicado en la plantilla.",
                value=f"{ccodcen} / {ctipcom}",
            )

        seen_keys.add(key)

        if not row_has_error:
            records.append(
                CaceteTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccodgen=ccodgen,
                    ccodcen=ccodcen,
                    cnomcen=cnomcen,
                    ctipcom=ctipcom,
                    cdescom=cdescom,
                    nvolalm=nvolalm,
                    nadqmes=nadqmes,
                )
            )

    expected_count = len(catalog.thermal_fuels)

    if len(records) != expected_count:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=None,
            field=None,
            message="La plantilla CACETE no contiene todos los combustibles térmicos esperados.",
            value=f"válidas={len(records)}; esperadas={expected_count}",
        )

    return CaceteTemplateValidationResult(
        template_path=template_path,
        period=period,
        records=records,
        issues=issues,
    )