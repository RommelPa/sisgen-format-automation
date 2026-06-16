from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.comcen.template import (
    COMCEN_HEADERS,
    DEFAULT_COMPANY_CODE,
    DEFAULT_FUEL_CODE,
    DEFAULT_FUEL_DESCRIPTION,
    CenterUnit,
    load_center_units,
    load_center_units_from_db,
    parse_period,
)


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class ComcenTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class ComcenTemplateRecord:
    canoreg: str
    cmesreg: str
    ccodemp: str
    ccodcen: str
    ctipgru: str
    cnomnum: str
    ccodcom: str
    cdescom: str
    ntotcom: Decimal


@dataclass(frozen=True)
class ComcenTemplateValidationResult:
    template_path: Path
    period: str
    records: list[ComcenTemplateRecord]
    issues: list[ComcenTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected = [header.upper() for header in COMCEN_HEADERS]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(COMCEN_HEADERS)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados COMCEN en la plantilla.")


def _catalog_key(unit: CenterUnit) -> tuple[str, str, str]:
    return (unit.ccodcen, unit.ctipgru, unit.cnomnum)


def validate_comcen_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path | None = None,
    catalog_db_path: Path | None = None,
) -> ComcenTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla COMCEN: {template_path}")

    if catalog_path is None and catalog_db_path is None:
        raise ValueError("Debes indicar catalog_path o catalog_db_path.")

    if catalog_path is not None and catalog_db_path is not None:
        raise ValueError("Use solo un origen de catalogo: YAML o SQLite.")

    expected_year, expected_month = parse_period(period)

    if catalog_db_path is not None:
        catalog_units = load_center_units_from_db(catalog_db_path)
    else:
        if catalog_path is None:
            raise ValueError("Debes indicar catalog_path.")
        catalog_units = load_center_units(catalog_path)

    expected_catalog_keys = {_catalog_key(unit) for unit in catalog_units}

    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows)

    issues: list[ComcenTemplateIssue] = []
    records: list[ComcenTemplateRecord] = []
    seen_keys: set[tuple[str, str, str, str, str, str, str]] = set()
    present_catalog_keys: set[tuple[str, str, str]] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(COMCEN_HEADERS)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(COMCEN_HEADERS, values, strict=True))

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccodemp = _normalize_text(raw["CCODEMP"])
        ccodcen = _normalize_text(raw["CCODCEN"])
        ctipgru = _normalize_text(raw["CTIPGRU"])
        cnomnum = _normalize_text(raw["CNOMNUM"])
        ccodcom = _normalize_text(raw["CCODCOM"])
        cdescom = _normalize_text(raw["CDESCOM"])
        ntotcom = _parse_decimal(raw["NTOTCOM"])

        row_has_error = False

        if canoreg != expected_year:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CANOREG",
                    message="CANOREG no coincide con el periodo.",
                    value=canoreg,
                )
            )

        if cmesreg != expected_month:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CMESREG",
                    message="CMESREG no coincide con el periodo.",
                    value=cmesreg,
                )
            )

        if ccodemp != DEFAULT_COMPANY_CODE:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODEMP",
                    message="CCODEMP fue modificado. Debe permanecer como EGASA.",
                    value=ccodemp,
                )
            )

        catalog_key = (ccodcen, ctipgru, cnomnum)

        if catalog_key not in expected_catalog_keys:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODCEN/CTIPGRU/CNOMNUM",
                    message="El grupo térmico no existe en el catálogo CENTER.",
                    value=" / ".join(catalog_key),
                )
            )
        else:
            present_catalog_keys.add(catalog_key)

        if ccodcom != DEFAULT_FUEL_CODE:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODCOM",
                    message="CCODCOM fue modificado. Debe permanecer como D2.",
                    value=ccodcom,
                )
            )

        if cdescom != DEFAULT_FUEL_DESCRIPTION:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CDESCOM",
                    message="CDESCOM fue modificado. Debe permanecer como Diesel 2       (Gl).",
                    value=cdescom,
                )
            )

        if ntotcom is None:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="NTOTCOM",
                    message="NTOTCOM es obligatorio. Usa 0 si no hubo consumo.",
                    value=raw["NTOTCOM"],
                )
            )
        elif ntotcom < Decimal("0"):
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="NTOTCOM",
                    message="NTOTCOM no puede ser negativo.",
                    value=ntotcom,
                )
            )

        unique_key = (
            canoreg,
            cmesreg,
            ccodemp,
            ccodcen,
            ctipgru,
            cnomnum,
            ccodcom,
        )

        if unique_key in seen_keys:
            row_has_error = True
            issues.append(
                ComcenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CLAVE",
                    message="Fila duplicada para el mismo grupo térmico y combustible.",
                    value=" / ".join(unique_key),
                )
            )

        seen_keys.add(unique_key)

        if not row_has_error and ntotcom is not None:
            records.append(
                ComcenTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccodemp=ccodemp,
                    ccodcen=ccodcen,
                    ctipgru=ctipgru,
                    cnomnum=cnomnum,
                    ccodcom=ccodcom,
                    cdescom=cdescom,
                    ntotcom=ntotcom,
                )
            )

    missing_catalog_keys = expected_catalog_keys - present_catalog_keys

    for missing_key in sorted(missing_catalog_keys):
        issues.append(
            ComcenTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=None,
                field="CATALOGO",
                message="Falta una fila COMCEN para un grupo térmico del catálogo CENTER.",
                value=" / ".join(missing_key),
            )
        )

    return ComcenTemplateValidationResult(
        template_path=template_path,
        period=period,
        records=records,
        issues=issues,
    )