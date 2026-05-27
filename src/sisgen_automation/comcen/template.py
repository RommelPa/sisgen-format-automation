from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

COMCEN_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODEMP",
    "CCODCEN",
    "CTIPGRU",
    "CNOMNUM",
    "CCODCOM",
    "CDESCOM",
    "NTOTCOM",
]

DEFAULT_COMPANY_CODE = "EGASA"
DEFAULT_FUEL_CODE = "D2"
DEFAULT_FUEL_DESCRIPTION = "Diesel 2       (Gl)"

EDITABLE_HEADERS = {"NTOTCOM"}


@dataclass(frozen=True)
class CenterUnit:
    central: str
    ccodcon: str
    ccodcen: str
    ctipgru: str
    cnomnum: str
    npotins: str
    npotefe: str


@dataclass(frozen=True)
class ComcenTemplateResult:
    period: str
    output_path: Path
    rows: int


def parse_period(period: str) -> tuple[str, str]:
    try:
        year_text, month_text = period.split("-", maxsplit=1)
    except ValueError as exc:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.") from exc

    if len(year_text) != 4 or len(month_text) != 2:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.")

    year = int(year_text)
    month = int(month_text)

    if not 1 <= month <= 12:
        raise ValueError("El mes del periodo debe estar entre 01 y 12.")

    return f"{year % 100:02d}", f"{month:02d}"


def load_center_units(catalog_path: Path) -> list[CenterUnit]:
    if not catalog_path.exists():
        raise ValueError(f"No existe el catálogo CENTER: {catalog_path}")

    data = yaml.safe_load(catalog_path.read_text(encoding="utf-8"))

    if not isinstance(data, dict):
        raise ValueError("El catálogo CENTER debe ser un YAML con estructura de diccionario.")

    units = data.get("units")

    if not isinstance(units, list):
        raise ValueError("El catálogo CENTER debe contener una lista 'units'.")

    parsed_units: list[CenterUnit] = []

    for index, raw_unit in enumerate(units, start=1):
        if not isinstance(raw_unit, dict):
            raise ValueError(f"La unidad CENTER #{index} no es un diccionario válido.")

        try:
            parsed_units.append(
                CenterUnit(
                    central=str(raw_unit["central"]).strip(),
                    ccodcon=str(raw_unit["ccodcon"]).strip(),
                    ccodcen=str(raw_unit["ccodcen"]).strip(),
                    ctipgru=str(raw_unit["ctipgru"]).strip(),
                    cnomnum=str(raw_unit["cnomnum"]).strip(),
                    npotins=str(raw_unit["npotins"]).strip(),
                    npotefe=str(raw_unit["npotefe"]).strip(),
                )
            )
        except KeyError as exc:
            raise ValueError(f"Falta el campo {exc} en la unidad CENTER #{index}.") from exc

    if not parsed_units:
        raise ValueError("El catálogo CENTER no contiene unidades.")

    return parsed_units


def default_comcen_template_path(period: str) -> Path:
    year_text, month_text = period.split("-", maxsplit=1)
    return Path("reports") / "templates" / f"COMCEN_{year_text}_{month_text}_template.xlsx"


def create_comcen_template(
    *,
    period: str,
    catalog_path: Path,
    output_path: Path | None = None,
) -> ComcenTemplateResult:
    year_short, month = parse_period(period)
    units = load_center_units(catalog_path)

    if output_path is None:
        output_path = default_comcen_template_path(period)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "COMCEN"

    header_fill = PatternFill("solid", fgColor="7030A0")   # morado encabezado
    locked_fill = PatternFill("solid", fgColor="DDEBF7")   # azul claro no editable
    editable_fill = PatternFill("solid", fgColor="FFF2CC") # amarillo editable
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 1

    header_fill = PatternFill("solid", fgColor="7030A0")
    locked_fill = PatternFill("solid", fgColor="DDEBF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_index, header in enumerate(COMCEN_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

    for row_index, unit in enumerate(units, start=header_row + 1):
        values: dict[str, Any] = {
            "CANOREG": year_short,
            "CMESREG": month,
            "CCODEMP": DEFAULT_COMPANY_CODE,
            "CCODCEN": unit.ccodcen,
            "CTIPGRU": unit.ctipgru,
            "CNOMNUM": unit.cnomnum,
            "CCODCOM": DEFAULT_FUEL_CODE,
            "CDESCOM": DEFAULT_FUEL_DESCRIPTION,
            "NTOTCOM": None,
        }

        for col_index, header in enumerate(COMCEN_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=values[header])

            if header in EDITABLE_HEADERS:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

            if header == "NTOTCOM":
                cell.number_format = "0.00000"

    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 18,
        "G": 10,
        "H": 24,
        "I": 14,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    for col_index in range(1, len(COMCEN_HEADERS) + 1):
        sheet.cell(row=header_row, column=col_index).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{header_row + len(units)}"
    sheet.protection.sheet = True
    sheet.protection.enable()

    workbook.save(output_path)

    return ComcenTemplateResult(
        period=period,
        output_path=output_path,
        rows=len(units),
    )