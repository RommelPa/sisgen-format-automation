from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.g7.catalog import load_g7_catalog, load_g7_catalog_from_db
from sisgen_automation.comene.template import ENERGY_HEADERS, parse_period



def _load_g7_validation_catalog(
    *,
    catalog_path: Path | None,
    catalog_db_path: Path | None,
):
    if catalog_db_path is not None:
        return load_g7_catalog_from_db(catalog_db_path)

    if catalog_path is not None:
        return load_g7_catalog(catalog_path)

    raise ValueError("Debe indicar catalog_path o catalog_db_path para G7.")

class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class EnergyTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class EnergyTemplateRecord:
    canoreg: str
    cmesreg: str
    ccosisi: str
    ccodeemp: str
    ccodgen: str
    cdesgen: str
    nenhopu: Decimal
    nenfuho: Decimal
    nenetot: Decimal
    nvahopu: Decimal
    nvafuho: Decimal
    nvaltot: Decimal


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, str) and value.strip().startswith("="):
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _find_header_row(rows: list[tuple[Any, ...]], headers: list[str]) -> int:
    expected = [header.upper() for header in headers]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(headers)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados en la plantilla.")


def _add_issue(
    issues: list[EnergyTemplateIssue],
    *,
    severity: IssueSeverity,
    row: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        EnergyTemplateIssue(
            severity=severity,
            row=row,
            field=field,
            message=message,
            value=value,
        )
    )


def _require_decimal(
    *,
    raw: dict[str, Any],
    field: str,
    row_index: int,
    issues: list[EnergyTemplateIssue],
) -> Decimal:
    parsed = _parse_decimal(raw[field])

    if parsed is None:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="Valor numérico obligatorio vacío o inválido.",
            value=raw[field],
        )
        return Decimal("0")

    if parsed < Decimal("0"):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="El valor numérico no puede ser negativo.",
            value=parsed,
        )

    return parsed


@dataclass(frozen=True)
class EnergyTemplateValidationResult:
    template_path: Path
    period: str
    records: list[EnergyTemplateRecord]
    issues: list[EnergyTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _validate_energy_template(
    *,
    template_path: Path,
    period: str,
    catalog,
    headers: list[str],
    expected_parties,
    parse_period_func,
) -> EnergyTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla: {template_path}")

    expected_year, expected_month = parse_period_func(period)
    expected_by_code = {
        item.ccodgen: item
        for item in expected_parties
    }

    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows, headers)

    issues: list[EnergyTemplateIssue] = []
    records: list[EnergyTemplateRecord] = []
    seen_keys: set[str] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(headers)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(headers, values, strict=True))
        initial_error_count = len(
            [issue for issue in issues if issue.severity == IssueSeverity.ERROR]
        )

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccosisi = _normalize_text(raw["CCOSISI"])
        ccodeemp = _normalize_text(raw["CCODEMP"])
        ccodgen = _normalize_text(raw["CCODGEN"])
        cdesgen = _normalize_text(raw["CDESGEN"])

        if canoreg != expected_year:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CANOREG",
                message="CANOREG no coincide con el periodo.",
                value=canoreg,
            )

        if cmesreg != expected_month:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CMESREG",
                message="CMESREG no coincide con el periodo.",
                value=cmesreg,
            )

        if ccosisi != catalog.system.ccosisi:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCOSISI",
                message="CCOSISI no coincide con el catálogo G7.",
                value=ccosisi,
            )

        if ccodeemp != catalog.company.ccodeemp:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODEMP",
                message="CCODEMP no coincide con el catálogo G7.",
                value=ccodeemp,
            )

        party = expected_by_code.get(ccodgen)

        if party is None:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODGEN",
                message="CCODGEN no existe en la sección esperada del catálogo G7.",
                value=ccodgen,
            )
        elif cdesgen != party.cdesgen:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CDESGEN",
                message="CDESGEN no coincide con el catálogo G7.",
                value=cdesgen,
            )

        nenhopu = _require_decimal(raw=raw, field="NENHOPU", row_index=row_index, issues=issues)
        nenfuho = _require_decimal(raw=raw, field="NENFUHO", row_index=row_index, issues=issues)
        nvahopu = _require_decimal(raw=raw, field="NVAHOPU", row_index=row_index, issues=issues)
        nvafuho = _require_decimal(raw=raw, field="NVAFUHO", row_index=row_index, issues=issues)

        nenetot = nenhopu + nenfuho
        nvaltot = nvahopu + nvafuho

        if ccodgen in seen_keys:
            _add_issue(
                issues,
                severity=IssueSeverity.ERROR,
                row=row_index,
                field="CCODGEN",
                message="CCODGEN duplicado en la plantilla.",
                value=ccodgen,
            )

        seen_keys.add(ccodgen)

        final_error_count = len(
            [issue for issue in issues if issue.severity == IssueSeverity.ERROR]
        )

        if final_error_count == initial_error_count:
            records.append(
                EnergyTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccosisi=ccosisi,
                    ccodeemp=ccodeemp,
                    ccodgen=ccodgen,
                    cdesgen=cdesgen,
                    nenhopu=nenhopu,
                    nenfuho=nenfuho,
                    nenetot=nenetot,
                    nvahopu=nvahopu,
                    nvafuho=nvafuho,
                    nvaltot=nvaltot,
                )
            )

    expected_keys = set(expected_by_code)

    for ccodgen in sorted(expected_keys - seen_keys):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=None,
            field="CCODGEN",
            message="Falta generador esperado según catálogo G7.",
            value=ccodgen,
        )

    for ccodgen in sorted(seen_keys - expected_keys):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=None,
            field="CCODGEN",
            message="Generador no esperado según catálogo G7.",
            value=ccodgen,
        )

    return EnergyTemplateValidationResult(
        template_path=template_path,
        period=period,
        records=records,
        issues=issues,
    )


def validate_comene_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path | None = None,
    catalog_db_path: Path | None = None,
) -> EnergyTemplateValidationResult:
    catalog = _load_g7_validation_catalog(
        catalog_path=catalog_path,
        catalog_db_path=catalog_db_path,
    )

    return _validate_energy_template(
        template_path=template_path,
        period=period,
        catalog=catalog,
        headers=ENERGY_HEADERS,
        expected_parties=catalog.energy_purchases,
        parse_period_func=parse_period,
    )
