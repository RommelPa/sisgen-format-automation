from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from sisgen_automation.g7.catalog import load_g7_catalog


@dataclass(frozen=True)
class ComeneTemplateResult:
    period: str
    output_path: Path
    rows: int

ENERGY_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCOSISI",
    "CCODEMP",
    "CCODGEN",
    "CDESGEN",
    "NENHOPU",
    "NENFUHO",
    "NENETOT",
    "NVAHOPU",
    "NVAFUHO",
    "NVALTOT",
]

EDITABLE_HEADERS = {
    "NENHOPU",
    "NENFUHO",
    "NVAHOPU",
    "NVAFUHO",
}

CALCULATED_HEADERS = {
    "NENETOT",
    "NVALTOT",
}

NUMERIC_HEADERS = {
    "NENHOPU",
    "NENFUHO",
    "NENETOT",
    "NVAHOPU",
    "NVAFUHO",
    "NVALTOT",
}


def parse_period(period: str) -> tuple[str, str]:
    try:
        year_text, month_text = period.split("-", maxsplit=1)
    except ValueError as exc:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.") from exc

    if len(year_text) != 4 or len(month_text) != 2:
        raise ValueError("El periodo debe tener formato YYYY-MM, por ejemplo 2026-01.")

    year = int(year_text)
    month = int(month_text)

    if not 1 <= month <= 12:
        raise ValueError("El mes del periodo debe estar entre 01 y 12.")

    return f"{year % 100:02d}", f"{month:02d}"


def _apply_sheet_style(
    *,
    sheet,
    row_count: int,
) -> None:
    widths = {
        "A": 10,
        "B": 10,
        "C": 10,
        "D": 12,
        "E": 12,
        "F": 45,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{1 + row_count}"
    sheet.sheet_view.showGridLines = False
    sheet.protection.sheet = True
    sheet.protection.enable()


def _create_energy_template(
    *,
    period: str,
    catalog_path: Path,
    output_path: Path,
    parties,
) -> int:
    year_short, month = parse_period(period)
    catalog = load_g7_catalog(catalog_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active

    header_fill = PatternFill("solid", fgColor="7030A0")
    locked_fill = PatternFill("solid", fgColor="DDEBF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    for col_index, header in enumerate(ENERGY_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

    for row_index, party in enumerate(parties, start=2):
        values = {
            "CANOREG": year_short,
            "CMESREG": month,
            "CCOSISI": catalog.system.ccosisi,
            "CCODEMP": catalog.company.ccodeemp,
            "CCODGEN": party.ccodgen,
            "CDESGEN": party.cdesgen,
            "NENHOPU": None,
            "NENFUHO": None,
            "NENETOT": f"=G{row_index}+H{row_index}",
            "NVAHOPU": None,
            "NVAFUHO": None,
            "NVALTOT": f"=J{row_index}+K{row_index}",
        }

        for col_index, header in enumerate(ENERGY_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=values[header])

            if header in EDITABLE_HEADERS:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

            if header in NUMERIC_HEADERS:
                cell.number_format = "0.00000"

    _apply_sheet_style(sheet=sheet, row_count=len(parties))

    workbook.save(output_path)

    return len(parties)


def default_comene_template_path(period: str) -> Path:
    year_text, month_text = period.split("-", maxsplit=1)
    return Path("reports") / "templates" / f"COMENE_{year_text}_{month_text}_template.xlsx"


def create_comene_template(
    *,
    period: str,
    catalog_path: Path,
    output_path: Path | None = None,
) -> ComeneTemplateResult:
    catalog = load_g7_catalog(catalog_path)

    if output_path is None:
        output_path = default_comene_template_path(period)

    rows = _create_energy_template(
        period=period,
        catalog_path=catalog_path,
        output_path=output_path,
        parties=catalog.energy_purchases,
    )

    return ComeneTemplateResult(
        period=period,
        output_path=output_path,
        rows=rows,
    )