from __future__ import annotations


from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sisgen_automation.g7.catalog import load_g7_catalog
from sisgen_automation.comnet.template import COMNET_HEADERS, parse_period


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class SimpleTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _find_header_row(rows: list[tuple[Any, ...]], headers: list[str]) -> int:
    expected = [header.upper() for header in headers]

    for index, row in enumerate(rows, start=1):
        values = [_normalize_text(value).upper() for value in row[: len(headers)]]

        if values == expected:
            return index

    raise ValueError("No se encontró la fila de encabezados en la plantilla.")


def _add_issue(
    issues: list[SimpleTemplateIssue],
    *,
    severity: IssueSeverity,
    row: int | None,
    field: str | None,
    message: str,
    value: object | None = None,
) -> None:
    issues.append(
        SimpleTemplateIssue(
            severity=severity,
            row=row,
            field=field,
            message=message,
            value=value,
        )
    )


def _require_decimal(
    *,
    raw: dict[str, Any],
    field: str,
    row_index: int,
    issues: list[SimpleTemplateIssue],
) -> Decimal:
    parsed = _parse_decimal(raw[field])

    if parsed is None:
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="Valor numérico obligatorio vacío o inválido.",
            value=raw[field],
        )
        return Decimal("0")

    if parsed < Decimal("0"):
        _add_issue(
            issues,
            severity=IssueSeverity.ERROR,
            row=row_index,
            field=field,
            message="El valor numérico no puede ser negativo.",
            value=parsed,
        )

    return parsed




@dataclass(frozen=True)
class ComnetTemplateRecord:
    canoreg: str
    cmesreg: str
    ccodeemp: str
    ccoddis: str
    cnomdis: str
    ncomnet: Decimal


@dataclass(frozen=True)
class ComnetTemplateValidationResult:
    template_path: Path
    period: str
    records: list[ComnetTemplateRecord]
    issues: list[SimpleTemplateIssue]

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def validate_comnet_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> ComnetTemplateValidationResult:
    if not template_path.exists():
        raise ValueError(f"No existe la plantilla COMNET: {template_path}")

    expected_year, expected_month = parse_period(period)
    catalog = load_g7_catalog(catalog_path)
    expected_by_code = {item.ccoddis: item for item in catalog.net_commitments}

    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows, COMNET_HEADERS)

    issues: list[SimpleTemplateIssue] = []
    records: list[ComnetTemplateRecord] = []
    seen_keys: set[str] = set()

    for row_index, row in enumerate(rows[header_row:], start=header_row + 1):
        values = list(row[: len(COMNET_HEADERS)])

        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw = dict(zip(COMNET_HEADERS, values, strict=True))
        initial_error_count = len([issue for issue in issues if issue.severity == IssueSeverity.ERROR])

        canoreg = _normalize_text(raw["CANOREG"])
        cmesreg = _normalize_text(raw["CMESREG"]).zfill(2)
        ccodeemp = _normalize_text(raw["CCODEMP"])
        ccoddis = _normalize_text(raw["CCODDIS"])
        cnomdis = _normalize_text(raw["CNOMDIS"])

        if canoreg != expected_year:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CANOREG", message="CANOREG no coincide con el periodo.", value=canoreg)

        if cmesreg != expected_month:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CMESREG", message="CMESREG no coincide con el periodo.", value=cmesreg)

        if ccodeemp != catalog.company.ccodeemp:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CCODEMP", message="CCODEMP no coincide con el catálogo G7.", value=ccodeemp)

        party = expected_by_code.get(ccoddis)

        if party is None:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CCODDIS", message="CCODDIS no existe en net_commitments del catálogo G7.", value=ccoddis)
        elif cnomdis != party.cnomdis:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CNOMDIS", message="CNOMDIS no coincide con el catálogo G7.", value=cnomdis)

        ncomnet = _require_decimal(raw=raw, field="NCOMNET", row_index=row_index, issues=issues)

        if ccoddis in seen_keys:
            _add_issue(issues, severity=IssueSeverity.ERROR, row=row_index, field="CCODDIS", message="CCODDIS duplicado en la plantilla.", value=ccoddis)

        seen_keys.add(ccoddis)

        final_error_count = len([issue for issue in issues if issue.severity == IssueSeverity.ERROR])

        if final_error_count == initial_error_count:
            records.append(
                ComnetTemplateRecord(
                    canoreg=canoreg,
                    cmesreg=cmesreg,
                    ccodeemp=ccodeemp,
                    ccoddis=ccoddis,
                    cnomdis=cnomdis,
                    ncomnet=ncomnet,
                )
            )

    expected_keys = set(expected_by_code)

    for ccoddis in sorted(expected_keys - seen_keys):
        _add_issue(issues, severity=IssueSeverity.ERROR, row=None, field="CCODDIS", message="Falta distribuidora esperada según catálogo G7.", value=ccoddis)

    for ccoddis in sorted(seen_keys - expected_keys):
        _add_issue(issues, severity=IssueSeverity.ERROR, row=None, field="CCODDIS", message="Distribuidora no esperada según catálogo G7.", value=ccoddis)

    return ComnetTemplateValidationResult(template_path=template_path, period=period, records=records, issues=issues)
