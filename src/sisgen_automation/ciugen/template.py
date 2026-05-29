from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from dbfread import DBF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from sisgen_automation.u2.catalog import load_u2_catalog
from sisgen_automation.u2.sources import DBF_ENCODING, parse_period


CIUGEN_HEADERS = [
    "CANOREG",
    "CMESREG",
    "CCODGEN",
    "CCODREG",
    "CCODDEP",
    "CCODCIU",
    "CDESCIU",
    "NUSULIB",
    "NCONLIB",
    "NFACTOT",
]

EDITABLE_HEADERS = {
    "NUSULIB",
    "NCONLIB",
    "NFACTOT",
}

NUMERIC_HEADERS = {
    "NUSULIB",
    "NCONLIB",
    "NFACTOT",
}


@dataclass(frozen=True)
class CiugenTemplateResult:
    period: str
    base_period: str
    output_path: Path
    rows: int


def default_ciugen_template_path(period: str) -> Path:
    year_text, month_text = period.split("-", maxsplit=1)
    return Path("reports") / "templates" / f"CIUGEN_{year_text}_{month_text}_template.xlsx"


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _full_year_from_sisgen_short_year(year_text: str) -> int:
    year_short = int(year_text)

    if year_short >= 90:
        return 1900 + year_short

    return 2000 + year_short


def _read_ciugen_rows(source_dbf_path: Path) -> list[dict[str, object]]:
    if not source_dbf_path.exists():
        raise ValueError(f"No existe CIUGEN.DBF fuente: {source_dbf_path}")

    table = DBF(
        str(source_dbf_path),
        load=True,
        encoding=DBF_ENCODING,
        char_decode_errors="ignore",
    )

    rows: list[dict[str, object]] = []

    for record in table:
        rows.append({header: record.get(header) for header in CIUGEN_HEADERS})

    return rows


def _latest_period(rows: list[dict[str, object]]) -> str:
    periods: set[tuple[int, int]] = set()

    for row in rows:
        year = _clean_text(row.get("CANOREG"))
        month = _clean_text(row.get("CMESREG")).zfill(2)

        if not year.isdigit() or not month.isdigit():
            continue

        month_int = int(month)

        if not 1 <= month_int <= 12:
            continue

        periods.add((_full_year_from_sisgen_short_year(year), month_int))

    if not periods:
        raise ValueError("No se encontraron periodos validos en CIUGEN.DBF.")

    year, month = max(periods)
    return f"{year:04d}-{month:02d}"


def _rows_for_period(rows: list[dict[str, object]], period: str) -> list[dict[str, object]]:
    year_short, month = parse_period(period)

    result: list[dict[str, object]] = []

    for row in rows:
        if _clean_text(row.get("CANOREG")) != year_short:
            continue

        if _clean_text(row.get("CMESREG")).zfill(2) != month:
            continue

        result.append(row)

    return result


def create_ciugen_template(
    *,
    period: str,
    source_dbf_path: Path,
    catalog_path: Path,
    base_period: str | None = None,
    output_path: Path | None = None,
) -> CiugenTemplateResult:
    year_short, month = parse_period(period)
    catalog = load_u2_catalog(catalog_path)

    rows = _read_ciugen_rows(source_dbf_path)

    if base_period is None:
        base_period = _latest_period(rows)

    base_rows = _rows_for_period(rows, base_period)

    if not base_rows:
        raise ValueError(f"No se encontraron registros CIUGEN para el periodo base {base_period}.")

    base_rows_by_code = {
        _clean_text(row.get("CCODCIU")): row
        for row in base_rows
    }

    if output_path is None:
        output_path = default_ciugen_template_path(period)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CIUGEN"

    header_fill = PatternFill("solid", fgColor="7030A0")
    locked_fill = PatternFill("solid", fgColor="DDEBF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    for col_index, header in enumerate(CIUGEN_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

    for row_index, item in enumerate(catalog.ciiu, start=2):
        base_row = base_rows_by_code.get(item.ccodciu)

        if base_row is None:
            # El catalogo manda. Si el CIIU no existia en el periodo base,
            # se crea con estructura valida y valores editables vacios.
            base_row = {}

        values: dict[str, Any] = {
            "CANOREG": year_short,
            "CMESREG": month,
            "CCODGEN": catalog.company.ccodgen,
            "CCODREG": catalog.location.ccodreg,
            "CCODDEP": catalog.location.ccoddep,
            "CCODCIU": item.ccodciu,
            "CDESCIU": item.description,
            "NUSULIB": None,
            "NCONLIB": None,
            "NFACTOT": None,
        }

        # Si en el futuro se decide precargar valores del periodo base,
        # hacerlo aqui de forma explicita. Por ahora los campos operativos
        # quedan vacios para obligar revision mensual.

        for col_index, header in enumerate(CIUGEN_HEADERS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=values[header])
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if header in EDITABLE_HEADERS:
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

            if header in NUMERIC_HEADERS:
                cell.number_format = "0.00000"

    widths = {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 64,
        "H": 14,
        "I": 16,
        "J": 16,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{1 + len(catalog.ciiu)}"
    sheet.sheet_view.showGridLines = False
    sheet.protection.sheet = True
    sheet.protection.enable()

    workbook.save(output_path)

    return CiugenTemplateResult(
        period=period,
        base_period=base_period,
        output_path=output_path,
        rows=len(catalog.ciiu),
    )
