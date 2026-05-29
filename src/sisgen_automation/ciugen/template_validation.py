from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

from sisgen_automation.ciugen.template import CIUGEN_HEADERS
from sisgen_automation.u2.catalog import load_u2_catalog
from sisgen_automation.u2.sources import parse_period


class IssueSeverity(str, Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass(frozen=True)
class CiugenTemplateIssue:
    severity: IssueSeverity
    row: int | None
    field: str | None
    message: str
    value: object | None = None


@dataclass(frozen=True)
class CiugenTemplateRecord:
    row: int
    values: dict[str, str | Decimal]


@dataclass(frozen=True)
class CiugenTemplateValidationResult:
    period: str
    records: tuple[CiugenTemplateRecord, ...]
    issues: tuple[CiugenTemplateIssue, ...]

    @property
    def errors(self) -> tuple[CiugenTemplateIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == IssueSeverity.ERROR)

    @property
    def warnings(self) -> tuple[CiugenTemplateIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == IssueSeverity.WARNING)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)

    @property
    def has_errors(self) -> bool:
        return self.error_count > 0


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _normalize_year(value: object) -> str:
    text = _clean_text(value)

    if text.endswith(".0"):
        text = text[:-2]

    return text.zfill(2)


def _normalize_month(value: object) -> str:
    text = _clean_text(value)

    if text.endswith(".0"):
        text = text[:-2]

    return text.zfill(2)


def _decimal_or_none(value: object) -> Decimal | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if text.startswith("="):
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _required_decimal(
    *,
    row: int,
    field: str,
    value: object,
    issues: list[CiugenTemplateIssue],
) -> Decimal:
    parsed = _decimal_or_none(value)

    if parsed is None:
        issues.append(
            CiugenTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=row,
                field=field,
                message="Valor numerico obligatorio vacio o invalido.",
                value=value,
            )
        )
        return Decimal("0")

    if parsed < 0:
        issues.append(
            CiugenTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=row,
                field=field,
                message="El valor no puede ser negativo.",
                value=value,
            )
        )

    return parsed


def _validate_headers(headers: list[str], issues: list[CiugenTemplateIssue]) -> None:
    if headers != CIUGEN_HEADERS:
        issues.append(
            CiugenTemplateIssue(
                severity=IssueSeverity.ERROR,
                row=1,
                field=None,
                message="La plantilla CIUGEN no tiene las columnas esperadas o estan en otro orden.",
                value=", ".join(headers),
            )
        )


def validate_ciugen_template(
    *,
    template_path: Path,
    period: str,
    catalog_path: Path,
) -> CiugenTemplateValidationResult:
    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla CIUGEN: {template_path}")

    year_short, month = parse_period(period)
    catalog = load_u2_catalog(catalog_path)
    ciiu_by_code = catalog.ciiu_by_code

    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook.active

    issues: list[CiugenTemplateIssue] = []
    headers = [_clean_text(sheet.cell(row=1, column=col).value) for col in range(1, len(CIUGEN_HEADERS) + 1)]
    _validate_headers(headers, issues)

    if issues:
        return CiugenTemplateValidationResult(period=period, records=(), issues=tuple(issues))

    records: list[CiugenTemplateRecord] = []
    seen_codes: set[str] = set()

    for row_index in range(2, sheet.max_row + 1):
        raw_values = {
            header: sheet.cell(row=row_index, column=col_index).value
            for col_index, header in enumerate(CIUGEN_HEADERS, start=1)
        }

        if all(value is None or str(value).strip() == "" for value in raw_values.values()):
            continue

        values: dict[str, str | Decimal] = {}

        canoreg = _normalize_year(raw_values["CANOREG"])
        cmesreg = _normalize_month(raw_values["CMESREG"])
        ccodgen = _clean_text(raw_values["CCODGEN"])
        ccodreg = _clean_text(raw_values["CCODREG"])
        ccoddep = _clean_text(raw_values["CCODDEP"])
        ccodciu = _clean_text(raw_values["CCODCIU"])
        cdesciu = _clean_text(raw_values["CDESCIU"])

        if canoreg != year_short:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CANOREG",
                    message=f"Anio invalido. Se esperaba {year_short}.",
                    value=raw_values["CANOREG"],
                )
            )

        if cmesreg != month:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CMESREG",
                    message=f"Mes invalido. Se esperaba {month}.",
                    value=raw_values["CMESREG"],
                )
            )

        if ccodgen != catalog.company.ccodgen:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODGEN",
                    message=f"Empresa invalida. Se esperaba {catalog.company.ccodgen}.",
                    value=ccodgen,
                )
            )

        if ccodreg != catalog.location.ccodreg:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODREG",
                    message=f"Region invalida. Se esperaba {catalog.location.ccodreg}.",
                    value=ccodreg,
                )
            )

        if ccoddep != catalog.location.ccoddep:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODDEP",
                    message=f"Departamento invalido. Se esperaba {catalog.location.ccoddep}.",
                    value=ccoddep,
                )
            )

        item = ciiu_by_code.get(ccodciu)
        if item is None:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="CCODCIU",
                    message="Codigo CIIU no existe en catalogo U2.",
                    value=ccodciu,
                )
            )
        else:
            if ccodciu in seen_codes:
                issues.append(
                    CiugenTemplateIssue(
                        severity=IssueSeverity.ERROR,
                        row=row_index,
                        field="CCODCIU",
                        message="Codigo CIIU duplicado en plantilla.",
                        value=ccodciu,
                    )
                )

            seen_codes.add(ccodciu)

            if cdesciu != item.description:
                issues.append(
                    CiugenTemplateIssue(
                        severity=IssueSeverity.ERROR,
                        row=row_index,
                        field="CDESCIU",
                        message="Descripcion CIIU no coincide con catalogo.",
                        value=cdesciu,
                    )
                )

        nusulib = _required_decimal(
            row=row_index,
            field="NUSULIB",
            value=raw_values["NUSULIB"],
            issues=issues,
        )
        nconlib = _required_decimal(
            row=row_index,
            field="NCONLIB",
            value=raw_values["NCONLIB"],
            issues=issues,
        )
        nfactot = _required_decimal(
            row=row_index,
            field="NFACTOT",
            value=raw_values["NFACTOT"],
            issues=issues,
        )

        if nusulib != nusulib.to_integral_value():
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=row_index,
                    field="NUSULIB",
                    message="NUSULIB debe ser entero.",
                    value=nusulib,
                )
            )

        values["CANOREG"] = year_short
        values["CMESREG"] = month
        values["CCODGEN"] = ccodgen
        values["CCODREG"] = ccodreg
        values["CCODDEP"] = ccoddep
        values["CCODCIU"] = ccodciu
        values["CDESCIU"] = cdesciu
        values["NUSULIB"] = nusulib
        values["NCONLIB"] = nconlib
        values["NFACTOT"] = nfactot

        records.append(CiugenTemplateRecord(row=row_index, values=values))

    for item in catalog.ciiu:
        if item.ccodciu not in seen_codes:
            issues.append(
                CiugenTemplateIssue(
                    severity=IssueSeverity.ERROR,
                    row=None,
                    field="CCODCIU",
                    message="Codigo CIIU del catalogo no aparece en la plantilla CIUGEN.",
                    value=f"{item.ccodciu} - {item.description}",
                )
            )

    valid_records = tuple(records) if not any(issue.severity == IssueSeverity.ERROR for issue in issues) else ()

    return CiugenTemplateValidationResult(
        period=period,
        records=valid_records,
        issues=tuple(issues),
    )
